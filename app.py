import io
import os
import re
import json
import math
import shutil
import hashlib
from decimal import Decimal
from datetime import datetime, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo
from collections import defaultdict, OrderedDict
from typing import Optional, Tuple, List, Dict

import pandas as pd
import streamlit as st
import openpyxl
import requests

# Google Drive upload (optional)
try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build as google_build
    from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
except Exception:
    service_account = None
    google_build = None
    MediaIoBaseUpload = None
    MediaIoBaseDownload = None

# -------------------- Optional deps --------------------
# Excel decrypt (SmartStore password 0000)
try:
    import msoffcrypto
except ModuleNotFoundError:
    msoffcrypto = None

# PDF screenshot render (PyMuPDF)
try:
    import fitz  # PyMuPDF (pymupdf)
except Exception:
    fitz = None

# Pillow (merge PNG pages -> one PNG)
try:
    from PIL import Image
except Exception:
    Image = None

# PDF text extract libs
try:
    import pdfplumber
except Exception:
    pdfplumber = None

try:
    from pypdf import PdfReader
except Exception:
    try:
        from PyPDF2 import PdfReader  # fallback
    except Exception:
        PdfReader = None

# ReportLab (PDFs)

# =====================================================
# Extracted modules
# =====================================================
from config import (
    # Timezone & datetime
    KST_TZ,
    KST,
    # Storage paths
    BACKUP_DIR,
    INVENTORY_FILE,
    EXPORT_ROOT,
    RULES_FILE,
    EXPR_RULES_PATH,
    MAPPING_PATH,
    STICKER_SETTINGS_PATH,
    TC_SETTINGS_PATH,
    TC_TEMPLATE_DEFAULT_PATH,
    # Constants
    COUNT_UNITS,
    EXCEL_PASSWORD,
    # Google Sync
    GOOGLE_SYNC_WEBAPP_URL,
    GOOGLE_SYNC_TOKEN,
    GOOGLE_SYNC_TIMEOUT_SEC,
    # ReportLab page sizes & layout
    A4,
    landscape,
    mm,
    colors,
    # ReportLab platypus
    SimpleDocTemplate,
    Table,
    LongTable,
    TableStyle,
    Paragraph,
    Spacer,
    KeepTogether,
    HRFlowable,
    # ReportLab styles & fonts
    ParagraphStyle,
    getSampleStyleSheet,
    TTFont,
    UnicodeCIDFont,
    pdfmetrics,
    canvas,
)
from drive_utils import save_excel_upload_to_drive_once
from sales import render_sales_calc_page
from invoice import render_invoice_register_page
from stock import render_bulk_stock_page


# -------------------- Atomic write helpers --------------------
def _atomic_write_text(path: str | Path, text: str, encoding: str = "utf-8") -> None:
    p = Path(path)
    tmp = p.with_suffix(p.suffix + ".tmp")
    tmp.write_text(text, encoding=encoding)
    tmp.replace(p)


def _atomic_write_bytes(path: str | Path, data: bytes) -> None:
    p = Path(path)
    tmp = p.with_suffix(p.suffix + ".tmp")
    tmp.write_bytes(data)
    tmp.replace(p)



def kst_date_folder() -> str:
    return datetime.now(KST).strftime("%Y.%m.%d")


def ensure_export_root() -> str:
    try:
        os.makedirs(EXPORT_ROOT, exist_ok=True)
    except Exception:
        pass
    return EXPORT_ROOT


def export_inventory_snapshot(df: pd.DataFrame) -> tuple[str, str]:
    """
    재고표(df)를 exports/YYYY.MM.DD/재고표_YYYY.MM.DD.xlsx 로 저장합니다.
    같은 날짜에 여러 번 내보내기를 누르면 파일은 덮어씁니다.
    """
    ensure_export_root()
    date_str = kst_date_folder()
    folder = os.path.join(EXPORT_ROOT, date_str)
    os.makedirs(folder, exist_ok=True)

    file_path = os.path.join(folder, f"재고표_{date_str}.xlsx")
    data = inventory_df_to_xlsx_bytes(df)
    with open(file_path, "wb") as f:
        f.write(data)
    return date_str, file_path


def list_export_dates() -> list[str]:
    ensure_export_root()
    try:
        names = os.listdir(EXPORT_ROOT)
    except Exception:
        return []

    out: list[str] = []
    for name in names:
        p = os.path.join(EXPORT_ROOT, name)
        if os.path.isdir(p) and re.fullmatch(r"\d{4}\.\d{2}\.\d{2}", name):
            out.append(name)

    out.sort(reverse=True)
    return out


def read_export_xlsx_bytes(date_str: str) -> bytes | None:
    p = os.path.join(EXPORT_ROOT, date_str, f"재고표_{date_str}.xlsx")
    if not os.path.exists(p):
        return None
    try:
        with open(p, "rb") as f:
            return f.read()
    except Exception:
        return None


def delete_export_date(date_str: str) -> bool:
    """exports/YYYY.MM.DD 폴더(해당 날짜 내보내기)를 통째로 삭제"""
    ensure_export_root()
    if not re.fullmatch(r"\d{4}\.\d{2}\.\d{2}", (date_str or "")):
        return False

    folder = os.path.join(EXPORT_ROOT, date_str)

    # 안전장치: exports 폴더 밖을 삭제하지 않도록 경로 검증
    root_abs = os.path.abspath(EXPORT_ROOT)
    folder_abs = os.path.abspath(folder)
    if not folder_abs.startswith(root_abs):
        return False

    if os.path.isdir(folder_abs):
        shutil.rmtree(folder_abs)
        return True
    return False


# =====================================================
# ✅ 고정품목 기능 사용 안 함 (현재 재고표 순서 그대로 유지)
# =====================================================
FIXED_PRODUCT_ORDER: list[str] = []


# =====================================================
# (A) PACK/BOX/EA 규칙 (1번코드)
# =====================================================
def norm_type(t: str) -> str:
    t = (t or "").strip()
    if t in ["팩", "PACK", "pack", "Pack"]:
        return "PACK"
    if t in ["박스", "BOX", "box", "Box"]:
        return "BOX"
    if t in ["개", "EA", "ea", "Each", "EACH"]:
        return "EA"
    return t.upper().strip()


def display_type(typ: str) -> str:
    typ = norm_type(typ)
    return {"PACK": "팩", "BOX": "박스", "EA": "개"}.get(typ, typ)


def parse_pack_size_g(val: str) -> float:
    """(PACK/EA) 값: 500 / 500g / 0.5kg 허용 -> g로 반환"""
    v = (val or "").strip().lower().replace(" ", "")
    if v.endswith("kg"):
        return float(v[:-2]) * 1000.0
    if v.endswith("g"):
        return float(v[:-1])
    return float(v)


def parse_box_size_kg(val: str) -> float:
    """(BOX) 값: 2 / 2kg / 2000g 허용 -> kg로 반환"""
    v = (val or "").strip().lower().replace(" ", "")
    if v.endswith("g"):
        return float(v[:-1]) / 1000.0
    if v.endswith("kg"):
        return float(v[:-2])
    return float(v)


def load_rules_text() -> str:
    if os.path.exists(RULES_FILE):
        try:
            with open(RULES_FILE, "r", encoding="utf-8") as f:
                return f.read()
        except Exception:
            pass

    return """# TYPE,상품명,값
# 팩(PACK),상품명,팩_기준_g(=1팩이 몇 g인지)  ex) 500 / 500g / 0.5kg
# 박스(BOX),상품명,박스_기준_kg(=1박스가 몇 kg인지) ex) 2 / 2kg / 2000g
# 개(EA),상품명,1개_기준_g(=1개가 몇 g인지) ex) 1kg / 500g
#
# ✅ 출력 규칙
# - 화면/결과는 모두 숫자만 출력(단위 글자 없음)
# - BOX 등록 상품은 1 미만이어도 나눠서 표시 (예: 600g / 2000g = 0.3)

# --- 팩 (자주 쓰는 것들) ---
팩,건대추,500
팩,방울토마토,500
팩,양송이,500
팩,완숙토마토,1kg

# --- 개 (자주 쓰는 것들) ---
개,깐마늘,1kg
개,청피망,500

# --- 박스 (총중량 ÷ 2kg => 박스) ---
박스,래디쉬,2
박스,적근대,2
박스,비타민,2
박스,쌈샐러리,2
박스,잎로메인,2
박스,적겨자,2
박스,적근대,2
박스,적치커리,2
박스,청치커리,2
박스,케일,2
박스,통로메인,2
박스,향나물,2
박스,뉴그린,2
박스,청경채,4
"""


def save_rules_text(text: str) -> None:
    _atomic_write_text(RULES_FILE, text or "", encoding="utf-8")


def parse_rules(text: str):
    pack_rules = {}  # {상품명: {"size_g": float}}
    box_rules = {}   # {상품명: {"size_kg": float}}
    ea_rules = {}    # {상품명: {"size_g": float}}

    for raw in (text or "").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue

        parts = [p.strip() for p in line.split(",")]
        if len(parts) < 3:
            continue

        typ = norm_type(parts[0])
        name = parts[1].strip()
        val_raw = parts[2].strip()

        try:
            if typ == "PACK":
                size_g = parse_pack_size_g(val_raw)
                if size_g > 0:
                    pack_rules[name] = {"size_g": size_g}

            elif typ == "BOX":
                size_kg = parse_box_size_kg(val_raw)
                if size_kg > 0:
                    box_rules[name] = {"size_kg": size_kg}

            elif typ == "EA":
                size_g = parse_pack_size_g(val_raw)
                if size_g > 0:
                    ea_rules[name] = {"size_g": size_g}
        except Exception:
            continue

    return pack_rules, box_rules, ea_rules


def upsert_rule(text: str, typ: str, name: str, val: str) -> str:
    typ_norm = norm_type(typ)
    typ_disp = display_type(typ_norm)

    name = (name or "").strip()
    val = (val or "").strip()
    if not typ_norm or not name or not val:
        return text

    lines = (text or "").splitlines()
    out = []
    replaced = False

    for ln in lines:
        if ln.strip().startswith("#") or not ln.strip():
            out.append(ln)
            continue

        parts = [p.strip() for p in ln.split(",")]
        if len(parts) >= 2 and norm_type(parts[0]) == typ_norm and parts[1] == name:
            out.append(f"{typ_disp},{name},{val}")
            replaced = True
        else:
            out.append(ln)

    if not replaced:
        if out and out[-1].strip() != "":
            out.append("")
        out.append(f"{typ_disp},{name},{val}")

    return "\n".join(out)


# =====================================================
# (B) 2번 코드: 매칭/표현 규칙 + 엑셀 처리 + PDF/TC 출력
# =====================================================
UNIT_PATTERNS = [
    r"\d+(?:\.\d+)?kg\s*~\s*\d+(?:\.\d+)?kg",
    r"\d+(?:\.\d+)?kg",
    r"(?:약\s*)?\d+(?:\.\d+)?g",
    r"\d+개",
    r"\d+통",
    r"\d+단",
    r"\d+봉",
    r"\d+팩",
]
UNIT_RE = re.compile(r"(" + "|".join(UNIT_PATTERNS) + r")")


def extract_variant(name: str) -> str:
    s = (name or "").strip()
    m = UNIT_RE.search(s)
    if not m:
        return ""
    u = m.group(0)
    u = re.sub(r"\s+", "", u)
    u = u.replace("약", "")
    if "~" in u:
        u = u.split("~", 1)[1]
    return u


def normalize_text(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _safe_int(v) -> Optional[int]:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return int(v)
    except Exception:
        return None


def _xml_escape(s: str) -> str:
    s = str(s or "")
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _text_width_pt(text: str, font: str, size: float) -> float:
    try:
        w = pdfmetrics.stringWidth(text, font, size)
        if not w or w <= 0:
            return len(text) * size * 0.55
        return w
    except Exception:
        return len(text) * size * 0.55


def fmt_qty(x):
    try:
        x = float(x)
        return int(x) if x.is_integer() else x
    except Exception:
        return x


def _as_int_qty(v) -> int:
    try:
        f = float(v)
        if abs(f - round(f)) < 1e-9:
            return int(round(f))
        return int(round(f))
    except Exception:
        return 0


# -------------------- TC defaults --------------------
TC_PRODUCT_NAME_FIXED = "채소팜상품"
TC_ACCESS_FALLBACK = "경비실 호출"
TC_TYPE_DAWN_DEFAULT = "자동"
TC_TYPE_NEXT_DEFAULT = "택배대행"

# 수취인별 PDF 스타일
RECIPIENT_FONT_SIZE = 12
RECIPIENT_LEADING = 15
RECIPIENT_BLOCK_GAP_MM = 4.0
RECIPIENT_LINE_AFTER_MM = 4.0

# 스티커 용지 설정 (21×29.5cm / 65칸 / 3.82×2.11cm)
STICKER_COLS = 5
STICKER_ROWS = 13
STICKER_PER_PAGE = STICKER_COLS * STICKER_ROWS  # 65

# 용지 크기 (mm)  -> 21×29.5cm
STICKER_PAGE_W_MM = 210.0
STICKER_PAGE_H_MM = 295.0

# 사용자 지정 여백 (cm -> mm)
STICKER_MARGIN_LEFT_MM = 4.0
STICKER_MARGIN_RIGHT_MM = 4.0
STICKER_MARGIN_TOP_MM = 11.0
STICKER_MARGIN_BOTTOM_MM = 10.0

# 스티커(라벨) 크기 (mm) -> 3.82×2.11cm
STICKER_CELL_W_MM = 38.2
STICKER_CELL_H_MM = 21.1

# 스티커 간격: 상/하 0cm, 좌/우 0.3cm
# ⚠️ 다만 "용지(21cm) - 여백(0.4cm*2)" 폭 안에 5칸을 맞추기 위해,
#     가로 간격은 필요 시 0.3cm보다 아주 조금(≈0.025cm) 줄어들 수 있습니다.
STICKER_GAP_X_MM = 3.0
STICKER_GAP_Y_MM = 0.0

# 글자
STICKER_FONT_SIZE = 13
STICKER_LEADING = 16

# 프린터 출력 보정(필요 시 수동 조정, 기본 0)
STICKER_OFFSET_X_MM = 0.0
STICKER_OFFSET_Y_MM = 0.0

# ✅ 행별 텍스트 위치 미세보정
# - 1~5행(상단 5줄): 상품명 위치를 "조금 더 위로"
# - 12~13행(하단 2줄): 상품명 위치를 "조금 더 아래로"
# (단위: mm, 필요하면 숫자만 조절하면 됩니다)
STICKER_TEXT_SHIFT_TOP_ROWS_MM = 3.5
STICKER_TEXT_SHIFT_BOTTOM_ROWS_MM = 3.5


def _clean_access_message(msg: str) -> str:
    s = str(msg or "").strip()
    return s if s else TC_ACCESS_FALLBACK


# -------------------- ✅ TC Settings (persist) --------------------
def load_tc_settings() -> Dict[str, str]:
    default = {"dawn": TC_TYPE_DAWN_DEFAULT, "next": TC_TYPE_NEXT_DEFAULT}
    if not TC_SETTINGS_PATH.exists():
        return default
    try:
        data = json.loads(TC_SETTINGS_PATH.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            return default
        dawn = normalize_text(data.get("dawn", "")) or TC_TYPE_DAWN_DEFAULT
        nxt = normalize_text(data.get("next", "")) or TC_TYPE_NEXT_DEFAULT
        return {"dawn": dawn, "next": nxt}
    except Exception:
        return default


def save_tc_settings(dawn: str, nxt: str) -> None:
    dawn = normalize_text(dawn) or TC_TYPE_DAWN_DEFAULT
    nxt = normalize_text(nxt) or TC_TYPE_NEXT_DEFAULT
    _atomic_write_text(
        TC_SETTINGS_PATH,
        json.dumps({"dawn": dawn, "next": nxt}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# -------------------- 표현규칙 (통/개/팩/봉 같은 단위 관리) --------------------
def default_expression_rules() -> Dict:
    return {
        "default_unit": "개",
        "units": [
            {"enabled": True, "unit": "개"},
            {"enabled": True, "unit": "봉"},
            {"enabled": True, "unit": "통"},
            {"enabled": True, "unit": "팩"},
            {"enabled": True, "unit": "단"},
        ],
        "note": "합산규칙(N)이 적용될 단위를 관리합니다.",
    }


def save_expression_rules(data: Dict) -> None:
    _atomic_write_text(EXPR_RULES_PATH, json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def load_expression_rules() -> Dict:
    if not EXPR_RULES_PATH.exists():
        data = default_expression_rules()
        save_expression_rules(data)
        return data
    try:
        data = json.loads(EXPR_RULES_PATH.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            raise ValueError("invalid")

        if "units" not in data or not isinstance(data["units"], list):
            data["units"] = default_expression_rules()["units"]

        if "default_unit" not in data or not isinstance(data["default_unit"], str):
            data["default_unit"] = default_expression_rules()["default_unit"]

        cleaned_units = []
        for r in data["units"]:
            u = normalize_text(r.get("unit", ""))
            if not u:
                continue
            cleaned_units.append({"enabled": bool(r.get("enabled", True)), "unit": u})
        data["units"] = cleaned_units

        # ✅ 기본 표현 단위에 "단"이 없으면 자동으로 추가 (기존 저장값 호환)
        migrated = False
        try:
            units_list = data.get("units", [])
            unit_names = [normalize_text(r.get("unit", "")) for r in units_list if isinstance(r, dict)]
            if "단" not in unit_names:
                # 되도록 "팩" 다음에 넣기
                if "팩" in unit_names:
                    pos = unit_names.index("팩") + 1
                    units_list.insert(pos, {"enabled": True, "unit": "단"})
                else:
                    units_list.append({"enabled": True, "unit": "단"})
                data["units"] = units_list
                migrated = True
        except Exception:
            migrated = False

        # 마이그레이션이 발생했으면 파일에도 반영
        if migrated:
            try:
                save_expression_rules(data)
            except Exception:
                pass
        data["default_unit"] = normalize_text(data.get("default_unit", "개")) or "개"
        return data
    except Exception:
        data = default_expression_rules()
        save_expression_rules(data)
        return data


def get_bundle_units(expr: Dict) -> List[str]:
    units = []
    for r in expr.get("units", []):
        if r.get("enabled", True):
            u = normalize_text(r.get("unit", ""))
            if u:
                units.append(u)

    seen = set()
    out = []
    for u in units:
        if u not in seen:
            out.append(u)
            seen.add(u)
    return out


def build_bundle_re(bundle_units: List[str]) -> re.Pattern:
    if not bundle_units:
        bundle_units = ["개"]
    unit_alt = "|".join(map(re.escape, bundle_units))
    return re.compile(rf"^\s*(\d+)\s*({unit_alt})\s*$")


# -------------------- 상품명 매칭 규칙 (합산규칙 N 포함) --------------------
def default_mapping_rules() -> List[Dict]:
    return [
        {
            "enabled": True,
            "match_type": "contains",
            "pattern": "와일드루꼴라",
            "display_name": "와일드",
            "sum_rule": None,
            "note": '예) "채소팜 와일드루꼴라 1kg ..." -> 와일드',
        },
        {
            "enabled": True,
            "match_type": "contains",
            "pattern": "라디치오",
            "display_name": "라디치오",
            "sum_rule": None,
            "note": '예) "채소팜 라디치오 1통 ..." -> 라디치오',
        },
        {
            "enabled": False,
            "match_type": "contains",
            "pattern": "오렌지",
            "display_name": "오렌지",
            "sum_rule": 5,
            "note": "예) 오렌지 합산규칙=5",
        },
    ]


def save_mapping_rules(rules: List[Dict]) -> None:
    _atomic_write_text(MAPPING_PATH, json.dumps(rules, ensure_ascii=False, indent=2), encoding="utf-8")


def load_mapping_rules() -> List[Dict]:
    if not MAPPING_PATH.exists():
        rules = default_mapping_rules()
        save_mapping_rules(rules)
        return rules

    try:
        raw = json.loads(MAPPING_PATH.read_text(encoding="utf-8"))
        if isinstance(raw, list):
            cleaned = []
            for r in raw:
                sr = _safe_int(r.get("sum_rule"))
                if sr is not None and sr < 2:
                    sr = None
                cleaned.append(
                    dict(
                        enabled=bool(r.get("enabled", True)),
                        match_type=normalize_text(r.get("match_type", "contains")) or "contains",
                        pattern=normalize_text(r.get("pattern", "")),
                        display_name=normalize_text(r.get("display_name", "")),
                        sum_rule=sr,
                        note=normalize_text(r.get("note", "")),
                    )
                )
            return cleaned
    except Exception:
        pass

    rules = default_mapping_rules()
    save_mapping_rules(rules)
    return rules


def apply_mapping(actual_name: str, rules: List[Dict]) -> Tuple[str, bool, Optional[int]]:
    actual = normalize_text(actual_name)
    if not actual:
        return "", False, None

    for r in rules:
        if not r.get("enabled", True):
            continue

        mt = normalize_text(r.get("match_type", "contains")) or "contains"
        pattern = normalize_text(r.get("pattern", ""))
        display = normalize_text(r.get("display_name", ""))
        sr = _safe_int(r.get("sum_rule"))

        if not pattern or not display:
            continue

        matched = False
        if mt == "exact":
            matched = (actual == pattern)
        elif mt == "contains":
            matched = (pattern in actual)
        elif mt == "regex":
            try:
                matched = bool(re.search(pattern, actual))
            except re.error:
                matched = False

        if matched:
            if sr is not None and sr < 2:
                sr = None
            return display, True, sr

    # fallback
    s = re.sub(r"^\s*채소팜\s*", "", actual)
    s = re.sub(r"\([^)]*\)", "", s).strip()
    s = re.sub(r"\s+", " ", s).strip()
    m = UNIT_RE.search(s)
    if m:
        s = s[: m.start()].strip()

    toks = s.split()
    if not toks:
        return actual, False, None

    PREFIX = {"생", "유기농", "국산", "수입", "냉동", "베이비", "프리미엄"}
    if len(toks) >= 2 and toks[0] in PREFIX:
        fallback = toks[0] + toks[1]
    else:
        fallback = toks[0]

    return fallback, False, None


def mapping_df_from_list(rules: List[Dict]) -> pd.DataFrame:
    df = pd.DataFrame(rules)
    keep = ["enabled", "match_type", "pattern", "display_name", "sum_rule", "note"]
    for c in keep:
        if c not in df.columns:
            df[c] = None
    return df[keep]


def mapping_list_from_df(edited: pd.DataFrame) -> List[Dict]:
    cleaned = []
    for _, row in edited.iterrows():
        pattern = normalize_text(row.get("pattern"))
        display = normalize_text(row.get("display_name"))
        if not pattern or not display:
            continue

        mt = normalize_text(row.get("match_type")) or "contains"
        if mt not in {"contains", "exact", "regex"}:
            mt = "contains"

        sr = _safe_int(row.get("sum_rule"))
        if sr is not None and sr < 2:
            sr = None

        cleaned.append(
            dict(
                enabled=bool(row.get("enabled", True)),
                match_type=mt,
                pattern=pattern,
                display_name=display,
                sum_rule=sr,
                note=normalize_text(row.get("note")),
            )
        )
    return cleaned


# -------------------- Backups (Excel) --------------------
def backup_rules_to_excel(mapping_rules: List[Dict], expr_rules: Dict) -> Path:
    out_path = BACKUP_DIR / "상품별매칭규칙_백업.xlsx"

    df_map = mapping_df_from_list(mapping_rules).rename(
        columns={
            "enabled": "사용",
            "match_type": "매칭방식",
            "pattern": "실제상품명(패턴)",
            "display_name": "표시될상품명",
            "sum_rule": "합산규칙(N)",
            "note": "메모",
        }
    )

    units = expr_rules.get("units", [])
    df_expr = pd.DataFrame(units)
    if df_expr.empty:
        df_expr = pd.DataFrame([{"enabled": True, "unit": expr_rules.get("default_unit", "개")}])
    if "enabled" not in df_expr.columns:
        df_expr["enabled"] = True
    if "unit" not in df_expr.columns:
        df_expr["unit"] = ""
    df_expr = df_expr[["enabled", "unit"]].rename(columns={"enabled": "사용", "unit": "단위"})

    df_meta = pd.DataFrame(
        [
            {"키": "default_unit", "값": expr_rules.get("default_unit", "개")},
            {"키": "note", "값": expr_rules.get("note", "")},
        ]
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_map.to_excel(writer, sheet_name="상품명매칭", index=False)
        df_expr.to_excel(writer, sheet_name="표현규칙_단위", index=False)
        df_meta.to_excel(writer, sheet_name="표현규칙_설정", index=False)

    return out_path


# -------------------- Sidebar panels (매칭 규칙 페이지에서만) --------------------
def sidebar_backup_folder():
    with st.sidebar.expander("📁 규칙 백업폴더", expanded=False):
        try:
            backups = sorted(BACKUP_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        except Exception:
            backups = []

        if not backups:
            st.caption("아직 백업 파일이 없습니다.")
            return

        for i, fp in enumerate(backups[:60]):
            cols = st.columns([6, 2, 2])
            cols[0].write(fp.name)

            try:
                b = fp.read_bytes()
                cols[1].download_button(
                    "다운",
                    data=b,
                    file_name=fp.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_bk_{i}_{fp.name}",
                    use_container_width=True,
                )
            except Exception:
                cols[1].write("")

            if cols[2].button("삭제", key=f"rm_bk_{i}_{fp.name}", use_container_width=True):
                try:
                    fp.unlink()
                    st.success(f"삭제 완료: {fp.name}")
                    st.rerun()
                except Exception:
                    st.error("삭제 실패")


def sidebar_expression_rules():
    expr = load_expression_rules()
    units = expr.get("units", [])
    default_unit = normalize_text(expr.get("default_unit", "개")) or "개"

    with st.sidebar.expander("⚙️ 표현규칙", expanded=False):
        st.caption("합산규칙(N)을 적용할 단위를 관리합니다. (통/개/팩/봉 등)")

        df = pd.DataFrame(units)
        if df.empty:
            df = pd.DataFrame([{"enabled": True, "unit": default_unit}])
        if "enabled" not in df.columns:
            df["enabled"] = True
        if "unit" not in df.columns:
            df["unit"] = ""
        df = df[["enabled", "unit"]]

        edited = st.data_editor(
            df,
            hide_index=True,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "enabled": st.column_config.CheckboxColumn("사용", default=True),
                "unit": st.column_config.TextColumn("단위"),
            },
            key="expr_units_editor",
        )

        enabled_units = []
        for _, r in edited.iterrows():
            u = normalize_text(r.get("unit", ""))
            if u:
                enabled_units.append((bool(r.get("enabled", True)), u))

        enabled_only = [u for en, u in enabled_units if en]
        if not enabled_only:
            enabled_only = ["개"]

        if default_unit not in enabled_only:
            default_unit = enabled_only[0]

        new_default = st.selectbox(
            "기본단위 (구분이 비어있을 때)",
            options=enabled_only,
            index=enabled_only.index(default_unit) if default_unit in enabled_only else 0,
            key="expr_default_unit",
        )

        if st.button("💾 표현규칙 저장", use_container_width=True, key="save_expr_rules_btn"):
            cleaned_units = []
            seen = set()
            for en, u in enabled_units:
                if u in seen:
                    continue
                cleaned_units.append({"enabled": bool(en), "unit": u})
                seen.add(u)

            data = {
                "default_unit": new_default,
                "units": cleaned_units,
                "note": expr.get("note", ""),
            }
            save_expression_rules(data)
            st.success("표현규칙 저장 완료")
            st.rerun()


# -------------------- Excel decrypt / read --------------------
def decrypt_excel(uploaded_bytes: bytes, password: str = EXCEL_PASSWORD) -> io.BytesIO:
    if msoffcrypto is None:
        raise ModuleNotFoundError("msoffcrypto not installed")
    decrypted = io.BytesIO()
    office = msoffcrypto.OfficeFile(io.BytesIO(uploaded_bytes))
    office.load_key(password=password)
    office.decrypt(decrypted)
    decrypted.seek(0)
    return decrypted


def _norm_col(x) -> str:
    s = str(x if x is not None else "")
    s = s.replace("\xa0", " ").replace("\n", " ").replace("\r", " ")
    return normalize_text(s)


def find_col(df: pd.DataFrame, keywords: List[str]) -> Optional[str]:
    """컬럼명 탐색: 공백/개행/NBSP 등을 정규화해서 매칭합니다."""
    cols = list(df.columns)
    if not cols:
        return None

    kw_norm = [_norm_col(k) for k in (keywords or []) if _norm_col(k)]
    if not kw_norm:
        return None

    col_norm = [_norm_col(c) for c in cols]

    # 1) 정규화 후 완전일치
    for k in kw_norm:
        for c, cn in zip(cols, col_norm):
            if k == cn:
                return c

    # 2) 정규화 후 부분일치
    for c, cn in zip(cols, col_norm):
        for k in kw_norm:
            if k in cn:
                return c

    return None


# -------------------- Smart Excel header detection --------------------
REQUIRED_COL_GROUPS = OrderedDict(
    [
        ("상품명", ["상품명", "상품", "제품명"]),
        ("수량", ["수량", "주문수량", "구매수량", "개수"]),
        ("구매자명", ["구매자명", "구매자"]),
        ("수취인명", ["수취인명", "수령인", "받는사람"]),
        ("통합배송지", ["통합배송지", "배송지", "주소"]),
        ("옵션정보", ["옵션정보", "옵션", "선택옵션"]),
        ("수취인연락처", ["수취인연락처", "수령인연락처", "수취인 연락처", "수령인 연락처", "전화번호", "연락처"]),
        ("배송메세지", ["배송메세지", "배송메시지", "배송 메시지", "배송 메세지", "배송요청사항", "요청사항"]),
    ]
)


# -------------------- ✅ Sticker Exclude Settings (persist) --------------------
def load_sticker_exclude() -> List[str]:
    """스티커용지 PDF에서 제외할 상품명 목록을 로드합니다."""
    if not STICKER_SETTINGS_PATH.exists():
        return []
    try:
        data = json.loads(STICKER_SETTINGS_PATH.read_text(encoding="utf-8"))
        # allow either list or {"exclude":[...]}
        if isinstance(data, dict):
            data = data.get("exclude", [])
        if not isinstance(data, list):
            return []
        out: List[str] = []
        seen = set()
        for x in data:
            s = normalize_text(x)
            if not s:
                continue
            if s in seen:
                continue
            out.append(s)
            seen.add(s)
        return out
    except Exception:
        return []


def save_sticker_exclude(exclude: List[str]) -> None:
    exclude = exclude or []
    out: List[str] = []
    seen = set()
    for x in exclude:
        s = normalize_text(x)
        if not s:
            continue
        if s in seen:
            continue
        out.append(s)
        seen.add(s)
    _atomic_write_text(
        STICKER_SETTINGS_PATH,
        json.dumps({"exclude": out}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )



def _missing_required_cols(df: pd.DataFrame) -> List[str]:
    missing = []
    for k, kws in REQUIRED_COL_GROUPS.items():
        if find_col(df, kws) is None:
            missing.append(k)
    return missing


def _guess_header_row(preview: pd.DataFrame, scan_limit: int = 40) -> Tuple[Optional[int], int]:
    """header=None 로 읽은 preview에서 '헤더로 보이는 행'을 추정"""
    if preview is None or preview.empty:
        return None, 0

    best_i = None
    best_score = -1

    n = min(scan_limit, len(preview))
    for i in range(n):
        row = preview.iloc[i].tolist()
        row_strs = []
        for v in row:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            s = _norm_col(v)
            if s:
                row_strs.append(s)

        if not row_strs:
            continue

        score = 0
        for _, kws in REQUIRED_COL_GROUPS.items():
            hit = False
            for kw in kws:
                kw_n = _norm_col(kw)
                if not kw_n:
                    continue
                if any(kw_n in cell for cell in row_strs):
                    hit = True
                    break
            if hit:
                score += 1

        if score > best_score:
            best_score = score
            best_i = i

    return best_i, max(best_score, 0)


def smart_read_orders_excel(excel_bytes: bytes, min_score: int = 4) -> Tuple[pd.DataFrame, Dict]:
    """
    스마트스토어/일괄발송 엑셀처럼 상단에 안내문이 있는 경우,
    '헤더 행'을 자동으로 찾아서 DataFrame을 반환합니다.
    """
    if not excel_bytes:
        raise ValueError("empty excel bytes")

    def _read(sheet, header, nrows=None):
        bio = io.BytesIO(excel_bytes)
        bio.seek(0)
        return pd.read_excel(bio, sheet_name=sheet, header=header, nrows=nrows, engine="openpyxl")

    bio0 = io.BytesIO(excel_bytes)
    bio0.seek(0)
    xls = pd.ExcelFile(bio0, engine="openpyxl")

    best_fallback = None  # (df, meta)

    for sheet in xls.sheet_names:
        # 1) 일반 header=0 시도
        try:
            df0 = _read(sheet, header=0)
            df0 = df0.dropna(how="all")
            missing0 = _missing_required_cols(df0)
            if not missing0:
                return df0, {"sheet": sheet, "header_row": 0, "method": "header=0"}
            if best_fallback is None:
                best_fallback = (df0, {"sheet": sheet, "header_row": 0, "method": "header=0", "missing": missing0})
        except Exception:
            pass

        # 2) header 행 추정
        try:
            preview = _read(sheet, header=None, nrows=60)
            header_row, score = _guess_header_row(preview, scan_limit=40)

            if header_row is None or score < min_score:
                continue

            df = _read(sheet, header=int(header_row))
            df = df.dropna(how="all")

            # 빈 컬럼 제거(Unnamed: n)
            try:
                df = df.loc[:, ~df.columns.astype(str).str.match(r"^Unnamed")]
            except Exception:
                pass

            missing = _missing_required_cols(df)
            if not missing:
                return df, {"sheet": sheet, "header_row": int(header_row), "method": f"guessed(score={score})"}

            if best_fallback is None:
                best_fallback = (
                    df,
                    {"sheet": sheet, "header_row": int(header_row), "method": f"guessed(score={score})", "missing": missing},
                )
        except Exception:
            pass

    if best_fallback is not None:
        return best_fallback

    df_last = _read(0, header=0)
    return df_last, {"sheet": 0, "header_row": 0, "method": "fallback"}


# -------------------- 합산규칙 적용 (표현규칙에서 켠 단위에만) --------------------
def parse_bundle_variant(variant: str, bundle_re: re.Pattern) -> Tuple[Optional[int], Optional[str]]:
    m = bundle_re.match((variant or "").strip())
    if not m:
        return None, None
    try:
        return int(m.group(1)), m.group(2)
    except Exception:
        return None, None


def explode_sum_rule_rows(
    df_rows: pd.DataFrame,
    bundle_units: List[str],
    default_unit: str,
) -> pd.DataFrame:
    bundle_units = bundle_units or [default_unit or "개"]
    default_unit = default_unit or "개"
    bundle_re = build_bundle_re(bundle_units)
    unit_set = set(bundle_units)

    out = []
    for _, r in df_rows.iterrows():
        product = r["제품명"]
        variant = (r.get("구분", "") or "").strip()
        qty = r.get("수량", None)
        rule_n = _safe_int(r.get("합산규칙", None))

        if rule_n is None or rule_n < 2:
            out.append({"제품명": product, "구분": variant, "수량": qty})
            continue

        # 구분이 비어 있으면 기본 단위 1개로 간주
        if variant == "":
            unit_size, unit_label = 1, default_unit
            is_bundle = unit_label in unit_set
        else:
            unit_size, unit_label = parse_bundle_variant(variant, bundle_re)
            is_bundle = (unit_size is not None and unit_label in unit_set)

        if not is_bundle:
            out.append({"제품명": product, "구분": variant, "수량": qty})
            continue

        try:
            total_units = int(round(float(qty))) * int(unit_size)
        except Exception:
            out.append({"제품명": product, "구분": variant, "수량": qty})
            continue

        if total_units <= 0:
            continue

        full = total_units // rule_n
        rem = total_units % rule_n

        if full > 0:
            out.append({"제품명": product, "구분": f"{rule_n}{unit_label}", "수량": full})
        if rem > 0:
            out.append({"제품명": product, "구분": f"{rem}{unit_label}", "수량": 1})

    return pd.DataFrame(out)


# -------------------- 배송 옵션 분류 & 그룹 규칙 (새벽 우선) --------------------
def classify_delivery(opt: str) -> str:
    s = str(opt or "")
    if "새벽배송" in s:
        return "새벽배송"
    if "익일배송" in s:
        return "익일배송"
    return "기타"


def decide_group_delivery(deliv_set: set) -> str:
    if "새벽배송" in deliv_set:
        return "새벽배송"
    if "익일배송" in deliv_set:
        return "익일배송"
    return "기타"


# -------------------- PDF 1) 제품별 개수 --------------------
def build_summary_pdf(summary_df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()

    font_name = "Helvetica"
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("HYGothic-Medium"))
        font_name = "HYGothic-Medium"
    except Exception:
        pass

    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title",
        parent=styles["Heading2"],
        fontName=font_name,
        fontSize=14,
        leading=18,
        spaceAfter=8,
    )

    elems = []
    elems.append(Paragraph("▣ 제품별 개수", title_style))
    elems.append(Spacer(1, 4))

    data = [["제품명", "구분", "수량"]]
    for _, row in summary_df.iterrows():
        data.append([str(row["제품명"]), str(row["구분"]), str(row["수량"])])

    table = LongTable(
        data,
        colWidths=[75 * mm, 60 * mm, 25 * mm],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("ALIGN", (2, 1), (2, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("TOPPADDING", (0, 0), (-1, 0), 6),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ("TOPPADDING", (0, 1), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
            ]
        )
    )

    elems.append(table)
    doc.build(elems)
    return buf.getvalue()


# -------------------- PDF 2) 수취인별 출력 --------------------
def build_recipient_pdf(entries: List[Dict[str, str]], footer_prefix: str = "") -> bytes:
    buf = io.BytesIO()

    font_name = "Helvetica"
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("HYGothic-Medium"))
        font_name = "HYGothic-Medium"
    except Exception:
        pass

    footer_prefix = (footer_prefix or "").strip()
    footer_font_size = 11

    left_margin = 12 * mm
    right_margin = 12 * mm
    top_margin = 12 * mm
    bottom_margin = 12 * mm

    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=top_margin,
        bottomMargin=bottom_margin,
    )

    styles = getSampleStyleSheet()
    base_style = ParagraphStyle(
        "base",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=RECIPIENT_FONT_SIZE,
        leading=RECIPIENT_LEADING,
        spaceAfter=0,
    )

    usable_width = A4[0] - left_margin - right_margin

    def _draw_footer(c: canvas.Canvas, _doc):
        # 하단 중앙 페이지 표기: "새벽 -1-" / "익일 -2-" ...
        if not footer_prefix:
            return
        try:
            page_no = int(c.getPageNumber())
        except Exception:
            page_no = 1

        txt = f"{footer_prefix} -{page_no}-"
        y = 6 * mm  # 하단 여백 안쪽에 고정

        c.saveState()
        try:
            c.setFont(font_name, footer_font_size)
        except Exception:
            c.setFont("Helvetica", footer_font_size)

        w = _text_width_pt(txt, font_name, footer_font_size)
        x = (A4[0] - w) / 2.0
        c.drawString(x, y, txt)
        c.restoreState()

    elems = []
    for e in entries:
        recv = (e.get("수취인명") or "").strip() or " "
        items = (e.get("items_line") or "").strip() or " "

        name_token_plain = f"{recv} - "
        indent = _text_width_pt(name_token_plain, font_name, base_style.fontSize)
        indent_cap = usable_width * 0.55
        indent = min(max(indent, 40), indent_cap)

        line_style = ParagraphStyle(
            f"line_{abs(hash(recv)) % 10_000_000}",
            parent=base_style,
            leftIndent=indent,
            firstLineIndent=-indent,
        )

        text = f"<b>{_xml_escape(recv)}</b> - {_xml_escape(items)}"
        p = Paragraph(text, line_style)

        block = KeepTogether(
            [
                p,
                Spacer(1, RECIPIENT_BLOCK_GAP_MM * mm),
                HRFlowable(
                    width="100%",
                    thickness=0.4,
                    color=colors.lightgrey,
                    spaceBefore=0,
                    spaceAfter=RECIPIENT_LINE_AFTER_MM * mm,
                ),
            ]
        )
        elems.append(block)

    doc.build(elems, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
    return buf.getvalue()


# -------------------- PDF 3) 스티커 용지 --------------------
def _wrap_for_cell(txt: str, font_name: str, font_size: int, max_w_pt: float) -> List[str]:
    txt = (txt or "").strip()
    if not txt:
        return [""]

    def w(s: str) -> float:
        return _text_width_pt(s, font_name, font_size)

    if w(txt) <= max_w_pt:
        return [txt]

    if " " in txt:
        parts = txt.split()
        line1 = ""
        consumed = 0
        for p in parts:
            cand = (line1 + " " + p).strip()
            if w(cand) <= max_w_pt:
                line1 = cand
                consumed += 1
            else:
                break
        rest = " ".join(parts[consumed:]).strip()
        if not rest:
            return [line1]
        if w(rest) <= max_w_pt:
            return [line1, rest]
        trimmed = rest
        while trimmed and w(trimmed + "...") > max_w_pt:
            trimmed = trimmed[:-1]
        return [line1, (trimmed + "...") if trimmed else "..."]

    line1 = ""
    for ch in txt:
        if w(line1 + ch) <= max_w_pt:
            line1 += ch
        else:
            break
    rest = txt[len(line1):].strip()
    if not rest:
        return [line1]
    if w(rest) <= max_w_pt:
        return [line1, rest]
    trimmed = rest
    while trimmed and w(trimmed + "...") > max_w_pt:
        trimmed = trimmed[:-1]
    return [line1, (trimmed + "...") if trimmed else "..."]


def _draw_center_text(c: canvas.Canvas, font_name: str, font_size: int, x_center: float, y: float, txt: str):
    txt = (txt or "").strip()
    if not txt:
        return
    w = _text_width_pt(txt, font_name, font_size)
    x_left = x_center - (w / 2.0)
    t = c.beginText()
    t.setTextOrigin(x_left, y)
    t.setFont(font_name, font_size)
    t.textOut(txt)
    c.drawText(t)


def build_sticker_pdf(label_texts: List[str]) -> bytes:
    """
    스티커(라벨) PDF 출력
    - 용지: 21×29.5cm
    - 여백: L/R 0.4cm, T 1.1cm, B 1.0cm
    - 라벨: 3.82×2.11cm, 5×13 = 65칸
    - 간격: 좌/우 0.3cm, 상/하 0cm (폭/여백을 맞추기 위해 가로 간격은 자동 보정될 수 있음)
    - 각 라벨 중앙에 상품명(텍스트) 출력
    """
    buf = io.BytesIO()

    font_name = "Helvetica"
    try:
        pdfmetrics.registerFont(UnicodeCIDFont("HYGothic-Medium"))
        font_name = "HYGothic-Medium"
    except Exception:
        pass

    pagesize = (STICKER_PAGE_W_MM * mm, STICKER_PAGE_H_MM * mm)
    c = canvas.Canvas(buf, pagesize=pagesize)
    page_w_pt, page_h_pt = pagesize

    left_pt = STICKER_MARGIN_LEFT_MM * mm
    right_pt = STICKER_MARGIN_RIGHT_MM * mm
    top_pt = STICKER_MARGIN_TOP_MM * mm
    bottom_pt = STICKER_MARGIN_BOTTOM_MM * mm

    cell_w_pt = STICKER_CELL_W_MM * mm
    cell_h_pt = STICKER_CELL_H_MM * mm

    # gap (가로는 "0.3cm" 목표이지만, 실제 폭/여백에 맞추기 위해 자동 보정)
    gap_x_target_pt = STICKER_GAP_X_MM * mm
    gap_y_target_pt = STICKER_GAP_Y_MM * mm

    usable_w = page_w_pt - left_pt - right_pt
    usable_h = page_h_pt - top_pt - bottom_pt

    # 가로 간격 자동 보정(그리드가 여백을 침범하면 gap을 줄여서 맞춤)
    if STICKER_COLS > 1:
        grid_w_target = (STICKER_COLS * cell_w_pt) + ((STICKER_COLS - 1) * gap_x_target_pt)
        if grid_w_target > usable_w + (0.1 * mm):
            gap_x_pt = max(0.0, (usable_w - (STICKER_COLS * cell_w_pt)) / (STICKER_COLS - 1))
        else:
            gap_x_pt = gap_x_target_pt
    else:
        gap_x_pt = 0.0

    # 세로는 기본 "0", 혹시라도 오차로 넘치면 gap을 줄여서(=0 유지) 맞춤
    if STICKER_ROWS > 1:
        grid_h_target = (STICKER_ROWS * cell_h_pt) + ((STICKER_ROWS - 1) * gap_y_target_pt)
        if grid_h_target > usable_h + (0.1 * mm):
            gap_y_pt = max(0.0, (usable_h - (STICKER_ROWS * cell_h_pt)) / (STICKER_ROWS - 1))
        else:
            gap_y_pt = gap_y_target_pt
    else:
        gap_y_pt = 0.0

    # 시작점: 좌측 여백 기준, 상단 여백 기준(ReportLab은 좌하단이 (0,0))
    x0 = left_pt + (STICKER_OFFSET_X_MM * mm)
    y_top = (page_h_pt - top_pt) + (STICKER_OFFSET_Y_MM * mm)  # 첫 줄 스티커의 윗변

    total = len(label_texts)
    page_count = (total + STICKER_PER_PAGE - 1) // STICKER_PER_PAGE if total else 1

    pad_x = 2.0 * mm
    max_text_w = cell_w_pt - (pad_x * 2)

    for p in range(page_count):
        c.setFillColor(colors.black)
        c.setFont(font_name, STICKER_FONT_SIZE)

        for r in range(STICKER_ROWS):
            for col in range(STICKER_COLS):
                slot = r * STICKER_COLS + col
                global_i = p * STICKER_PER_PAGE + slot
                if global_i >= total:
                    continue

                text = (label_texts[global_i] or "").strip()
                if not text:
                    continue

                x = x0 + col * (cell_w_pt + gap_x_pt)
                y = y_top - ((r + 1) * cell_h_pt) - (r * gap_y_pt)  # 셀의 하단

                # ✅ 행별 텍스트 위치 미세보정 (요청: 1~5행 ↑ / 12~13행 ↓)
                row_shift_pt = 0.0
                if 0 <= r <= 4:
                    row_shift_pt += STICKER_TEXT_SHIFT_TOP_ROWS_MM * mm
                elif r in (11, 12):
                    row_shift_pt -= STICKER_TEXT_SHIFT_BOTTOM_ROWS_MM * mm

                lines = _wrap_for_cell(text, font_name, STICKER_FONT_SIZE, max_text_w)[:2]
                cx = x + cell_w_pt / 2.0

                if len(lines) == 1:
                    cy = y + (cell_h_pt / 2.0) - (STICKER_FONT_SIZE * 0.35) + row_shift_pt
                    _draw_center_text(c, font_name, STICKER_FONT_SIZE, cx, cy, lines[0])
                else:
                    center = y + (cell_h_pt / 2.0) + row_shift_pt
                    upper_y = center + (STICKER_LEADING * 0.25)
                    lower_y = center - (STICKER_LEADING * 0.95)
                    _draw_center_text(c, font_name, STICKER_FONT_SIZE, cx, upper_y, lines[0])
                    _draw_center_text(c, font_name, STICKER_FONT_SIZE, cx, lower_y, lines[1])

        if p < page_count - 1:
            c.showPage()

    c.save()
    return buf.getvalue()



# -------------------- TC 주문_등록양식 자동 채우기 --------------------
def _norm_header(s: str) -> str:
    s = str(s or "")
    s = s.replace("*", "")
    s = re.sub(r"\s+", "", s)
    return s.strip().lower()


def build_tc_excel_bytes(template_bytes: bytes, rows: List[Dict[str, str]]) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    if "양식" not in wb.sheetnames:
        raise ValueError("TC 템플릿에 '양식' 시트가 없습니다.")
    ws = wb["양식"]

    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        if v is None:
            continue
        headers[_norm_header(v)] = col

    def col_of(label_candidates: List[str]) -> int:
        for cand in label_candidates:
            key = _norm_header(cand)
            if key in headers:
                return headers[key]
        raise KeyError(f"필수 헤더를 찾지 못했습니다: {label_candidates}")

    c_req = col_of(["배송요청일", "배송요청일*"])
    c_orderer = col_of(["주문자", "주문자*"])
    c_receiver = col_of(["수령자", "수령자*"])
    c_addr = col_of(["수령자도로명주소", "수령자 도로명 주소", "수령자 도로명 주소*"])
    c_phone = col_of(["수령자연락처", "수령자 연락처", "수령자 연락처*"])
    c_in = col_of(["출입방법", "출입 방법"])
    c_prod = col_of(["상품명", "상품명*"])
    c_type = col_of(["배송유형", "배송 유형", "배송 유형*"])

    start_row = 2
    for i, r in enumerate(rows):
        rr = start_row + i
        ws.cell(rr, c_req).value = r.get("배송요청일", "")
        ws.cell(rr, c_orderer).value = r.get("주문자", "")
        ws.cell(rr, c_receiver).value = r.get("수령자", "")
        ws.cell(rr, c_addr).value = r.get("수령자도로명주소", "")
        ws.cell(rr, c_phone).value = r.get("수령자연락처", "")
        ws.cell(rr, c_in).value = r.get("출입방법", "")
        ws.cell(rr, c_prod).value = r.get("상품명", "")
        ws.cell(rr, c_type).value = r.get("배송유형", "")
        # 배송받을장소는 건드리지 않음

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =====================================================
# (C) 1번 코드: PDF(스크린샷/합계표) + 재고관리
# =====================================================
def render_pdf_pages_to_images(file_bytes: bytes, zoom: float = 2.0) -> list[bytes]:
    """
    PDF 각 페이지를 PNG 스크린샷으로 렌더링하여 bytes 리스트 반환
    zoom: 1.0~3.5 (클수록 선명/용량 증가)
    """
    if fitz is None:
        raise RuntimeError("스크린샷 저장은 pymupdf가 필요합니다. (pip install pymupdf)")

    doc = fitz.open(stream=file_bytes, filetype="pdf")
    out: list[bytes] = []
    mat = fitz.Matrix(zoom, zoom)

    for i in range(doc.page_count):
        page = doc.load_page(i)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        out.append(pix.tobytes("png"))

    doc.close()
    return out


def merge_png_pages_to_one(png_bytes_list: list[bytes]) -> bytes:
    """
    여러 PNG(페이지)를 세로로 이어붙여 1장 PNG로 반환
    Pillow(PIL) 필요
    """
    if not png_bytes_list:
        return b""

    if len(png_bytes_list) == 1:
        return png_bytes_list[0]

    if Image is None:
        return png_bytes_list[0]

    imgs = [Image.open(io.BytesIO(b)).convert("RGBA") for b in png_bytes_list]
    max_w = max(im.width for im in imgs)
    total_h = sum(im.height for im in imgs)

    canvas_img = Image.new("RGBA", (max_w, total_h), (255, 255, 255, 0))
    y = 0
    for im in imgs:
        x = (max_w - im.width) // 2
        canvas_img.paste(im, (x, y))
        y += im.height

    out = io.BytesIO()
    canvas_img.save(out, format="PNG", optimize=True)
    return out.getvalue()


def fmt_num(x: float, max_dec=2) -> str:
    s = f"{x:.{max_dec}f}".rstrip("0").rstrip(".")
    return s if s else "0"


def format_weight(grams: float) -> str | None:
    """kg/g도 숫자만: kg 소수로 표시 (19kg250g -> 19.25)"""
    if grams <= 0:
        return None
    kg = grams / 1000.0
    return fmt_num(kg, 3)


def parse_spec_components(spec: str):
    if not spec:
        return None

    s = spec.replace(",", "").replace(" ", "")
    s = s.replace("㎏", "kg").replace("ＫＧ", "kg").replace("KG", "kg").lower()

    out = {"grams_per_unit": None, "bunch_per_unit": None, "counts_per_unit": {}}

    # ✅ 19kg250g 같은 결합 표기 지원
    m2 = re.search(r"(\d+(?:\.\d+)?)kg(\d+(?:\.\d+)?)g", s)
    if m2:
        kg = float(m2.group(1))
        g = float(m2.group(2))
        out["grams_per_unit"] = kg * 1000.0 + g
    else:
        mw = re.search(r"(\d+(?:\.\d+)?)(kg|g)", s)
        if mw:
            num = float(mw.group(1))
            unit = mw.group(2)
            out["grams_per_unit"] = num * 1000.0 if unit == "kg" else num

    mb = re.search(r"(\d+)단", s)
    if mb:
        out["bunch_per_unit"] = int(mb.group(1))

    for u in COUNT_UNITS:
        mu = re.search(r"(\d+)" + re.escape(u), s)
        if mu:
            out["counts_per_unit"][u] = int(mu.group(1))

    if out["grams_per_unit"] is None and out["bunch_per_unit"] is None and not out["counts_per_unit"]:
        return None
    return out


def aggregate(items: list[tuple[str, str, float]]):
    agg = defaultdict(lambda: {"grams": 0.0, "bunch": 0.0, "counts": defaultdict(float), "unknown": defaultdict(float)})

    for product, spec, qty in items:
        try:
            q = float(qty)
        except Exception:
            q = 0.0

        comp = parse_spec_components(spec)
        if comp is None:
            agg[product]["unknown"][spec] += q
            continue

        if comp["grams_per_unit"] is not None:
            agg[product]["grams"] += float(comp["grams_per_unit"]) * q

        if comp["bunch_per_unit"] is not None:
            agg[product]["bunch"] += float(comp["bunch_per_unit"]) * q

        for unit, n in comp["counts_per_unit"].items():
            agg[product]["counts"][unit] += float(n) * q

    return agg


def _append_count_parts(parts: list[str], counts: dict):
    for u in ["개", "팩", "통", "봉"]:
        v = counts.get(u, 0)
        if v:
            # 소수는 거의 없겠지만 혹시 있으면 그대로
            if abs(v - round(v)) < 1e-9:
                parts.append(f"{int(round(v))}")
            else:
                parts.append(fmt_num(float(v), 2))


def format_total_custom(product: str, rec, pack_rules, box_rules, ea_rules,
                        allow_decimal_pack: bool, allow_decimal_box: bool) -> str:
    parts: list[str] = []

    # 단도 숫자만
    if rec["bunch"]:
        if abs(rec["bunch"] - round(rec["bunch"])) < 1e-9:
            parts.append(f'{int(round(rec["bunch"]))}')
        else:
            parts.append(fmt_num(float(rec["bunch"]), 2))

    grams = float(rec["grams"])
    counts = dict(rec["counts"])

    # BOX 우선: 박스 기준으로 나눈 값(0.3처럼) 표시 (1 미만이어도 항상 표시)
    if product in box_rules and grams > 0:
        box_size_kg = float(box_rules[product]["size_kg"])
        denom_g = box_size_kg * 1000.0
        boxes = grams / denom_g

        if allow_decimal_box:
            parts.append(f"{fmt_num(boxes, 2)}")
        else:
            if abs(boxes - round(boxes)) < 1e-9:
                parts.append(f"{int(round(boxes))}")
            else:
                parts.append(f"{fmt_num(boxes, 2)}")

        _append_count_parts(parts, counts)
        return " ".join(parts).strip() if parts else "0"

    # PACK / EA 처리
    pack_shown = False
    ea_shown = False

    # spec 자체에 팩이 있으면 우선
    if counts.get("팩", 0) > 0:
        v = counts.get("팩", 0)
        parts.append(f"{int(round(v))}" if abs(v - round(v)) < 1e-9 else fmt_num(float(v), 2))
        pack_shown = True
        counts.pop("팩", None)

    # rules로 g -> 팩 변환
    elif product in pack_rules and grams > 0:
        size_g = float(pack_rules[product]["size_g"])
        packs = grams / size_g
        if allow_decimal_pack:
            parts.append(f"{fmt_num(packs, 2)}")
            pack_shown = True
        else:
            if abs(packs - round(packs)) < 1e-9:
                parts.append(f"{int(round(packs))}")
                pack_shown = True

    # 팩이 안 잡혔으면 "개" 처리
    if not pack_shown:
        if counts.get("개", 0) > 0:
            v = counts.get("개", 0)
            parts.append(f"{int(round(v))}" if abs(v - round(v)) < 1e-9 else fmt_num(float(v), 2))
            ea_shown = True
            counts.pop("개", None)

        elif product in ea_rules and grams > 0:
            size_g = float(ea_rules[product]["size_g"])
            eas = grams / size_g
            # 정수로 딱 떨어질 때만 표시(아니면 중량 kg 소수로)
            if abs(eas - round(eas)) < 1e-9:
                parts.append(f"{int(round(eas))}")
                ea_shown = True

    # 팩도 개도 안 잡히면 중량(kg 소수)
    if not pack_shown and not ea_shown:
        w = format_weight(grams)
        if w:
            parts.append(w)

    _append_count_parts(parts, counts)
    return " ".join(parts).strip() if parts else "0"


def to_3_per_row(df: pd.DataFrame, n: int = 3) -> pd.DataFrame:
    """
    ✅ 세로 우선 배치(위→아래), 그 다음 열로 이동
    n=3이면 1열을 위→아래로 다 채운 뒤 2열, 3열 순서
    """
    if df is None or len(df) == 0:
        row = {}
        for c in range(n):
            row[f"제품명{c+1}"] = ""
            row[f"합계{c+1}"] = ""
        return pd.DataFrame([row])

    total = len(df)
    rows_count = math.ceil(total / n)

    out = []
    for r in range(rows_count):
        row = {}
        for c in range(n):
            idx = c * rows_count + r  # 세로 우선
            if idx < total:
                row[f"제품명{c+1}"] = df.iloc[idx]["제품명"]
                row[f"합계{c+1}"] = df.iloc[idx]["합계"]
            else:
                row[f"제품명{c+1}"] = ""
                row[f"합계{c+1}"] = ""
        out.append(row)

    return pd.DataFrame(out)


def make_pdf_bytes(df: pd.DataFrame, title: str) -> bytes:
    """
    1번 코드 스타일(landscape A4 + NanumGothic 폰트) 유지
    """
    font_path = os.path.join("fonts", "NanumGothic.ttf")
    font_name = "NanumGothic"

    if not os.path.exists(font_path):
        raise RuntimeError(f"폰트 파일을 못 찾음: {font_path} (fonts 폴더/파일명 확인)")

    if font_name not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(font_name, font_path))
        pdfmetrics.registerFontFamily(
            font_name, normal=font_name, bold=font_name, italic=font_name, boldItalic=font_name
        )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18
    )

    styles = getSampleStyleSheet()
    title_style = styles["Title"].clone("KTitle")
    title_style.fontName = font_name

    cell_style = ParagraphStyle(
        "KCell", fontName=font_name, fontSize=10, leading=12,
        alignment=1, wordWrap="CJK"
    )
    header_style = ParagraphStyle(
        "KHeader", fontName=font_name, fontSize=10, leading=12,
        alignment=1, wordWrap="CJK"
    )

    elements = [Paragraph(title, title_style), Spacer(1, 12)]
    safe_df = df.fillna("").astype(str)

    header = [Paragraph(str(c), header_style) for c in safe_df.columns]
    body = [[Paragraph(str(v), cell_style) for v in row] for row in safe_df.values.tolist()]
    data = [header] + body

    page_w, _ = landscape(A4)
    usable_w = page_w - 36
    col_w = usable_w / max(1, len(safe_df.columns))
    col_widths = [col_w] * len(safe_df.columns)

    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


# =====================================================
# 재고관리 (1번 코드)
# =====================================================
INVENTORY_COLUMNS = [
    "상품명",
    "재고",
    "입고",
    "보유수량",
    "1차",
    "2차",
    "3차",
    "주문수량",
    "남은수량",
]


def _coerce_num_series(s: pd.Series) -> pd.Series:
    """숫자/소수 허용 (빈값/문자 -> 0)"""
    return pd.to_numeric(s, errors="coerce").fillna(0.0).astype(float)


def compute_inventory_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    if "상품명" not in df.columns:
        df.insert(0, "상품명", "")

    for col in ["재고", "입고", "1차", "2차", "3차"]:
        if col not in df.columns:
            df[col] = 0

    for col in ["재고", "입고", "1차", "2차", "3차"]:
        df[col] = _coerce_num_series(df[col])

    df["상품명"] = df["상품명"].fillna("").astype(str).str.strip()

    def _to_decimal(v):
        if v is None:
            return Decimal("0")
        try:
            if isinstance(v, float) and math.isnan(v):
                return Decimal("0")
            return Decimal(str(v))
        except Exception:
            return Decimal("0")

    stock_dec = [_to_decimal(v) for v in df["재고"].tolist()]
    in_dec = [_to_decimal(v) for v in df["입고"].tolist()]
    one_dec = [_to_decimal(v) for v in df["1차"].tolist()]
    two_dec = [_to_decimal(v) for v in df["2차"].tolist()]
    three_dec = [_to_decimal(v) for v in df["3차"].tolist()]

    have_dec = [a + b for a, b in zip(stock_dec, in_dec)]
    order_dec = [a + b + c for a, b, c in zip(one_dec, two_dec, three_dec)]
    remain_dec = [a - b for a, b in zip(have_dec, order_dec)]

    df["보유수량"] = [float(x) for x in have_dec]
    df["주문수량"] = [float(x) for x in order_dec]
    df["남은수량"] = [float(x) for x in remain_dec]

    for c in ["보유수량", "주문수량", "남은수량"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
        df[c] = df[c].mask(df[c].abs() < 1e-12, 0.0)

    return df[INVENTORY_COLUMNS]


def sort_inventory_df(df: pd.DataFrame) -> pd.DataFrame:
    """고정정렬 없이 현재 행 순서를 그대로 유지"""
    df = df.copy()
    for c in INVENTORY_COLUMNS:
        if c not in df.columns:
            df[c] = "" if c == "상품명" else 0.0
    return df[INVENTORY_COLUMNS].reset_index(drop=True)


def load_inventory_df() -> pd.DataFrame:
    if os.path.exists(INVENTORY_FILE):
        try:
            df = pd.read_csv(INVENTORY_FILE, encoding="utf-8-sig")
        except Exception:
            df = pd.read_csv(INVENTORY_FILE, encoding="utf-8", errors="ignore")
    else:
        df = pd.DataFrame(columns=["상품명", "재고", "입고", "1차", "2차", "3차"])

    df = compute_inventory_df(df)
    df = sort_inventory_df(df)
    df = df[df["상품명"].astype(str).str.strip() != ""].reset_index(drop=True)
    return df


def save_inventory_df(df: pd.DataFrame) -> None:
    p = Path(INVENTORY_FILE)
    tmp = p.with_suffix(p.suffix + ".tmp")
    df.to_csv(tmp, index=False, encoding="utf-8-sig")
    tmp.replace(p)


def parse_sum_to_number(total_str: str) -> float:
    """제품별합계 '합계' 문자열에서 첫 번째 숫자만 뽑아 등록용 수치로 사용"""
    s = (total_str or "").strip()
    nums = re.findall(r"[-+]?\d*\.?\d+", s)
    if not nums:
        return 0.0
    try:
        return float(nums[0])
    except Exception:
        return 0.0


def build_sync_items_from_sum(sum_df_long: pd.DataFrame) -> list[dict]:
    items: list[dict] = []
    if sum_df_long is None or len(sum_df_long) == 0:
        return items

    for _, r in sum_df_long.iterrows():
        name = str(r.get("제품명", "")).strip()
        if not name:
            continue
        qty = parse_sum_to_number(str(r.get("합계", "0")))
        items.append({"name": name, "qty": float(qty)})

    return items


def get_today_sheet_name_mdd() -> str:
    now = datetime.now(KST_TZ)
    return f"{now.month}.{now.day:02d}"


def sync_inventory_to_google_sheet(sum_df_long: pd.DataFrame, target_col: str, add_mode: bool = False) -> tuple[bool, str]:
    if not GOOGLE_SYNC_WEBAPP_URL or not GOOGLE_SYNC_TOKEN:
        return False, "구글시트 동기화 설정이 없어 로컬 재고표만 저장했습니다."

    payload = {
        "token": GOOGLE_SYNC_TOKEN,
        "sheet_name": get_today_sheet_name_mdd(),
        "target_col": target_col,
        "add_mode": bool(add_mode),
        "clear_target_first": not bool(add_mode),
        "source": "stock-app",
        "sent_at": datetime.now(KST_TZ).isoformat(),
        "items": build_sync_items_from_sum(sum_df_long),
    }

    try:
        resp = requests.post(GOOGLE_SYNC_WEBAPP_URL, json=payload, timeout=GOOGLE_SYNC_TIMEOUT_SEC)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        return False, f"구글시트 동기화 실패: {e}"

    if not bool(data.get("ok", False)):
        msg = str(data.get("message", "동기화 실패"))
        return False, f"구글시트 동기화 실패: {msg}"

    if not bool(data.get("sheet_found", True)):
        return True, f"구글시트 날짜 시트({payload['sheet_name']})가 없어 시트에는 반영하지 않았습니다."

    updated = int(data.get("updated", 0) or 0)
    not_found = data.get("not_found_names", []) or []
    if not_found:
        return True, f"구글시트 반영 {updated}건, 제외 {len(not_found)}건: {', '.join(map(str, not_found))}"
    return True, f"구글시트 반영 완료! (반영 행: {updated})"


def zero_numeric_inventory_fields(df: pd.DataFrame) -> pd.DataFrame:
    dd = df.copy()
    for c in ["재고", "입고", "1차", "2차", "3차"]:
        if c not in dd.columns:
            dd[c] = 0.0
        dd[c] = 0.0
    return dd


def register_sum_to_inventory(sum_df_long: pd.DataFrame, target_col: str, add_mode: bool = False):
    """제품별합계(df_long)를 재고관리의 1차/2차/3차 중 하나로 등록(상품명이 있는 것만)"""
    if sum_df_long is None or len(sum_df_long) == 0:
        return 0, []

    if "inventory_df" in st.session_state:
        inv = st.session_state["inventory_df"].copy()
    else:
        inv = load_inventory_df()

    inv = compute_inventory_df(inv)
    inv_names = inv["상품명"].fillna("").astype(str).str.strip()
    name_to_idx = {n: i for i, n in enumerate(inv_names) if n}

    skipped = []
    updated = 0

    if not add_mode and target_col in inv.columns:
        inv[target_col] = 0.0

    for _, r in sum_df_long.iterrows():
        name = str(r.get("제품명", "")).strip()
        if not name:
            continue
        if name not in name_to_idx:
            skipped.append(name)
            continue

        qty = parse_sum_to_number(str(r.get("합계", "0")))
        i = name_to_idx[name]

        if add_mode:
            inv.at[i, target_col] = float(inv.at[i, target_col]) + float(qty)
        else:
            inv.at[i, target_col] = float(qty)

        updated += 1

    inv = compute_inventory_df(inv)
    inv = sort_inventory_df(inv).reset_index(drop=True)

    st.session_state["inventory_df"] = inv
    save_inventory_df(inv)

    return updated, skipped


def inventory_df_to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="재고표")
        ws = writer.sheets["재고표"]
        ws.freeze_panes = "B2"
        widths = {"A": 16, "B": 8, "C": 8, "D": 10, "E": 8, "F": 8, "G": 8, "H": 10, "I": 10}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
    return buf.getvalue()


# =====================================================
# ✅ (핵심 변경) 엑셀 요약(summary_df) → 1번 제품별합계(합계) 계산
# =====================================================
def summary_to_items(summary_df: pd.DataFrame, default_unit: str) -> list[tuple[str, str, float]]:
    """
    2번 코드 결과(제품명/구분/수량)를 1번 코드 aggregate() 입력 형태로 변환.
    - 구분이 비어있으면 default_unit(기본단위)로 1개 처리: "1개" 같은 spec 생성
    """
    items: list[tuple[str, str, float]] = []
    if summary_df is None or len(summary_df) == 0:
        return items

    default_unit = normalize_text(default_unit) or "개"

    for _, r in summary_df.iterrows():
        product = str(r.get("제품명", "")).strip()
        if not product:
            continue

        spec = str(r.get("구분", "") or "").strip()
        if spec.lower() in ("nan", "none"):
            spec = ""
        if spec in ("", "-"):
            spec = f"1{default_unit}"

        try:
            qty = float(r.get("수량", 0) if r.get("수량", 0) is not None else 0)
        except Exception:
            qty = 0.0

        if qty == 0:
            continue

        items.append((product, spec, qty))

    return items


def compute_product_totals_from_summary(
    summary_df: pd.DataFrame,
    pack_rules,
    box_rules,
    ea_rules,
    allow_decimal_pack: bool,
    allow_decimal_box: bool,
    default_unit: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    반환: (df_long[제품명, 합계], df_wide[3열 배치])
    """
    items = summary_to_items(summary_df, default_unit=default_unit)
    agg = aggregate(items)

    rows = []
    for product in agg.keys():
        rows.append({
            "제품명": product,
            "합계": format_total_custom(
                product, agg[product],
                pack_rules, box_rules, ea_rules,
                allow_decimal_pack=allow_decimal_pack,
                allow_decimal_box=allow_decimal_box,
            ),
        })

    df_long = pd.DataFrame(rows)
    df_wide = to_3_per_row(df_long, 3)
    return df_long, df_wide


# =====================================================
# Streamlit UI (1번 코드 레이아웃 유지)
# =====================================================
st.set_page_config(
    page_title="재고프로그램",
    page_icon="assets/favicon.png",  # ✅ 1번 코드 파비콘/디자인 유지
    layout="wide",
)

# ----- Navigation -----
if "page" not in st.session_state:
    # ✅ 요청: 처음 열면 "엑셀 업로드"가 먼저
    st.session_state["page"] = "excel_results"

with st.sidebar:
    st.markdown("## 📌 메뉴")
    if st.button("📥 엑셀 업로드", use_container_width=True):
        st.session_state["page"] = "excel_results"
        st.rerun()
    if st.button("🧾 제품별 합계", use_container_width=True):
        st.session_state["page"] = "product_totals"
        st.rerun()
    if st.button("📦 재고관리", use_container_width=True):
        st.session_state["page"] = "inventory"
        st.rerun()
    if st.button("🚚 송장등록", use_container_width=True):
        st.session_state["page"] = "invoice_register"
        st.rerun()
    if st.button("🧰 재고일괄변경", use_container_width=True):
        st.session_state["page"] = "bulk_stock"
        st.rerun()
    if st.button("💰 매출계산", use_container_width=True):
        st.session_state["page"] = "sales_calc"
        st.rerun()
    if st.button("🧩 상품명 매칭 규칙", use_container_width=True):
        st.session_state["page"] = "mapping_rules"
        st.rerun()
    st.divider()


# =====================================================
# Pages
# =====================================================
def render_mapping_rules_page():
    # 🔒 비밀번호 보호 (상품명 매칭 규칙)
    if "mapping_authed" not in st.session_state:
        st.session_state["mapping_authed"] = False

    if not st.session_state["mapping_authed"]:
        st.title("🔒 상품명 매칭 규칙")
        st.caption("이 페이지는 비밀번호가 필요합니다.")

        pw = None
        ok = False
        left_col, _ = st.columns([2, 8])
        with left_col:
            with st.form("mapping_pw_form", clear_on_submit=False):
                pw = st.text_input(
                    "비밀번호",
                    type="password",
                    label_visibility="collapsed",
                    placeholder="비밀번호",
                )
                ok = st.form_submit_button("입장", use_container_width=False)

        if ok:
            if (pw or "").strip() == "1390":
                st.session_state["mapping_authed"] = True
                st.success("인증 완료!")
                st.rerun()
            else:
                st.error("비밀번호가 올바르지 않습니다.")
        return

    st.title("🧩 상품명 매칭 규칙")
    if st.button("🔓 로그아웃", use_container_width=False, key="mapping_logout_btn"):
        st.session_state["mapping_authed"] = False
        st.success("잠금 상태로 전환되었습니다.")
        st.rerun()

    st.caption("엑셀의 실제 상품명 → 표시될 상품명으로 매핑하고, 합산규칙(N)도 설정합니다.")

    sidebar_backup_folder()
    sidebar_expression_rules()

    mapping_rules = load_mapping_rules()
    expr = load_expression_rules()

    st.markdown(
        """
**매칭방식 설명**
- **contains**: `패턴`이 `엑셀 상품명` 안에 포함되면 매칭
- **exact**: `패턴`과 `엑셀 상품명`이 완전히 동일할 때만 매칭
- **regex**: `패턴`을 정규식으로 해석해 매칭

**합산규칙(N)**  
- N=5, 단위가 표현규칙에 포함된 경우(개/봉/통/팩 등) → 8개 주문 시 `5개 1개` + `3개 1개`로 표현
"""
    )

    df = mapping_df_from_list(mapping_rules)
    edited = st.data_editor(
        df,
        use_container_width=True,
        num_rows="dynamic",
        hide_index=True,
        column_config={
            "enabled": st.column_config.CheckboxColumn("사용", default=True),
            "match_type": st.column_config.SelectboxColumn("매칭 방식", options=["contains", "exact", "regex"]),
            "pattern": st.column_config.TextColumn("실제 상품명(패턴)", width="large"),
            "display_name": st.column_config.TextColumn("표시될 상품명", width="medium"),
            "sum_rule": st.column_config.NumberColumn("합산규칙(N)", min_value=2, step=1),
            "note": st.column_config.TextColumn("메모", width="large"),
        },
        key="mapping_editor_main",
    )

    c1, c2 = st.columns([1, 1])
    with c1:
        if st.button("💾 저장", use_container_width=True):
            cleaned = mapping_list_from_df(edited)
            save_mapping_rules(cleaned)
            st.success(f"저장 완료! (규칙 {len(cleaned)}개)")
            st.rerun()

    with c2:
        if st.button("📗 엑셀로 저장하기(백업)", use_container_width=True):
            cleaned_map = mapping_list_from_df(edited)
            outp = backup_rules_to_excel(cleaned_map, expr)
            st.success(f"백업 저장 완료: {outp.name}")
            st.rerun()


def render_excel_results_page():
    st.title("📥 엑셀 업로드")
    st.caption("엑셀 업로드 → 제품별 집계 + 수취인별 PDF + 스티커용지 PDF + TC주문_등록양식 자동작성")
    st.markdown("---")

    if msoffcrypto is None:
        st.error("msoffcrypto가 설치되지 않았습니다. requirements.txt에 'msoffcrypto-tool'을 추가하고 재배포해 주세요.")
        st.stop()

    # ✅ 이 페이지의 사이드바에서만 TC 배송유형 설정 + 저장 (요청사항 유지)
    tc_saved = load_tc_settings()
    if "tc_type_dawn" not in st.session_state:
        st.session_state.tc_type_dawn = tc_saved["dawn"]
    if "tc_type_next" not in st.session_state:
        st.session_state.tc_type_next = tc_saved["next"]

    with st.sidebar.expander("🔧 배송방법 설정", expanded=False):
        st.caption("변경 후 [저장]을 누르면 다음 실행에도 그대로 유지됩니다.")

        dawn_val = st.text_input(
            "새벽배송 → 배송유형",
            value=st.session_state.tc_type_dawn,
            key="tc_type_dawn_input",
        )

        next_val = st.text_input(
            "익일배송 → 배송유형",
            value=st.session_state.tc_type_next,
            key="tc_type_next_input",
        )

        if st.button("💾 TC 설정 저장", use_container_width=True, key="save_tc_settings_btn"):
            dawn_val = (dawn_val or "").strip() or TC_TYPE_DAWN_DEFAULT
            next_val = (next_val or "").strip() or TC_TYPE_NEXT_DEFAULT

            st.session_state.tc_type_dawn = dawn_val
            st.session_state.tc_type_next = next_val

            save_tc_settings(dawn_val, next_val)
            st.success("TC 설정 저장 완료")
            st.rerun()

    save_to_drive_choice = st.radio(
        "업로드한 엑셀을 Google Drive에 저장할까요?",
        options=["예", "아니오"],
        index=None,
        horizontal=True,
        key="orders_excel_save_to_drive_choice",
        help="예를 직접 선택한 경우에만 판매내역/올해/오늘날짜 폴더에 저장합니다. 연도 폴더가 없으면 자동 생성합니다. 아무것도 선택하지 않으면 저장하지 않습니다.",
    )

    uploaded = st.file_uploader("비밀번호(0000) 엑셀 업로드 (.xlsx)", type=["xlsx"], key="orders_excel_uploader")
    if uploaded is None:
        st.info("엑셀을 업로드하면 결과 표와 다운로드가 나타납니다.")
        st.stop()

    uploaded_original_bytes = uploaded.getvalue()

    # ✅ 사용자가 라디오 버튼에서 '예'를 직접 선택한 경우에만 Google Drive에 저장/덮어쓰기
    if save_to_drive_choice == "예":
        drive_ok, drive_msg = save_excel_upload_to_drive_once(getattr(uploaded, "name", "업로드엑셀.xlsx"), uploaded_original_bytes, target_dt=datetime.now(KST_TZ))
        if drive_ok:
            st.caption(f"✅ {drive_msg}")
        else:
            # Drive 저장 실패가 기존 엑셀 처리 기능을 막지 않도록 경고만 표시하고 계속 진행합니다.
            st.warning(drive_msg)
    elif save_to_drive_choice == "아니오":
        st.caption("Google Drive 저장 안 함으로 선택했습니다.")
    else:
        st.caption("Google Drive 저장 여부를 선택하지 않아 Drive에는 저장하지 않습니다.")

    upload_day = datetime.now(KST_TZ).date()
    req_day = upload_day + timedelta(days=1)
    req_day_str = req_day.strftime("%Y-%m-%d")

    try:
        decrypted_io = decrypt_excel(uploaded_original_bytes, password=EXCEL_PASSWORD)
        excel_bytes = decrypted_io.getvalue()
        raw_df, read_meta = smart_read_orders_excel(excel_bytes)
    except Exception as e:
        st.error('엑셀 읽기/복호화 실패: 비밀번호 "0000" 또는 파일 형식을 확인해 주세요.')
        st.exception(e)
        st.stop()

    if isinstance(read_meta, dict) and read_meta.get("method") != "header=0":
        st.caption(
            f"📌 헤더 자동탐지: sheet={read_meta.get('sheet')} / header_row={read_meta.get('header_row')} / {read_meta.get('method')}"
        )

    col_name = find_col(raw_df, ["상품명", "상품", "제품명"])
    col_qty = find_col(raw_df, ["수량", "주문수량", "구매수량", "개수"])
    col_buyer = find_col(raw_df, ["구매자명", "구매자"])
    col_recv = find_col(raw_df, ["수취인명", "수령인", "받는사람"])
    col_addr = find_col(raw_df, ["통합배송지", "배송지", "주소"])
    col_opt = find_col(raw_df, ["옵션정보", "옵션", "선택옵션"])
    col_recv_phone = find_col(raw_df, ["수취인연락처", "수령인연락처", "수취인 연락처", "수령인 연락처", "전화번호", "연락처"])
    col_msg = find_col(raw_df, ["배송메세지", "배송메시지", "배송 메시지", "배송 메세지", "배송요청사항", "요청사항"])

    missing = [k for k, v in {
        "상품명": col_name,
        "수량": col_qty,
        "구매자명": col_buyer,
        "수취인명": col_recv,
        "통합배송지": col_addr,
        "옵션정보": col_opt,
        "수취인연락처": col_recv_phone,
        "배송메세지": col_msg,
    }.items() if v is None]
    if missing:
        st.error(f"필수 컬럼을 찾지 못했습니다: {', '.join(missing)}")
        st.write("현재 컬럼:", list(raw_df.columns))
        st.stop()

    mapping_rules = load_mapping_rules()
    expr = load_expression_rules()
    bundle_units = get_bundle_units(expr)
    default_unit = normalize_text(expr.get("default_unit", "개")) or "개"

    work = raw_df[[col_buyer, col_recv, col_addr, col_recv_phone, col_msg, col_opt, col_name, col_qty]].copy()
    work.columns = ["구매자명", "수취인명", "통합배송지", "수취인연락처", "배송메세지", "옵션정보", "상품명", "수량"]

    work["상품명"] = work["상품명"].astype(str)
    work["수량"] = pd.to_numeric(work["수량"], errors="coerce")
    work["구분"] = work["상품명"].apply(extract_variant)

    mapped = work["상품명"].apply(lambda x: apply_mapping(x, mapping_rules))
    work["제품명"] = mapped.apply(lambda t: t[0])
    work["매칭성공"] = mapped.apply(lambda t: t[1])
    work["합산규칙"] = mapped.apply(lambda t: t[2])

    base = work[(work["수량"].notna()) & (work["제품명"] != "")].copy()

    exploded = explode_sum_rule_rows(
        base[["제품명", "구분", "수량", "합산규칙"]],
        bundle_units=bundle_units,
        default_unit=default_unit,
    )

    summary = (
        exploded.groupby(["제품명", "구분"], as_index=False)["수량"]
        .sum()
        .sort_values(["제품명", "구분"], kind="mergesort")
        .reset_index(drop=True)
    )
    summary["수량"] = summary["수량"].apply(fmt_qty)

    # ✅ 다른 페이지(제품별 합계)에서 바로 쓰도록 저장
    st.session_state["excel_summary_df"] = summary.copy()
    st.session_state["excel_default_unit"] = default_unit

    # -------------------- Results --------------------
    with st.expander("✅ 결과 (제품명 / 구분 / 수량)", expanded=False):
        st.dataframe(summary, use_container_width=True, height=520)

    with st.expander("⚠️ 미매칭/누락 행 (규칙 추가용)", expanded=False):
        bad = work[(work["매칭성공"] == False) | (work["수량"].isna())].copy()
        st.dataframe(bad.head(300), use_container_width=True)

    # 제품별 개수 PDF 다운로드
    st.download_button(
        "⬇️ 제품별 개수 PDF 다운로드",
        data=build_summary_pdf(summary),
        file_name="제품별개수.pdf",
        mime="application/pdf",
        use_container_width=True,
    )

    # 스티커 PDF
    st.markdown("---")
    st.subheader("🏷️ 스티커용지 PDF")

    # ✅ 스티커로 출력하지 않을 상품 설정 (펼쳐보기)
    # - 저장한 제외목록은 data/sticker_settings.json 에 남아 이후에도 자동 적용됩니다.
    if "sticker_exclude_products" not in st.session_state:
        st.session_state["sticker_exclude_products"] = load_sticker_exclude()

    product_options = sorted(
        [p for p in summary["제품명"].dropna().astype(str).str.strip().unique().tolist() if p]
    )

    saved_all = st.session_state.get("sticker_exclude_products", []) or []
    saved_in_options = [p for p in saved_all if p in product_options]
    saved_outside = [p for p in saved_all if p not in product_options]

    # 기본은 "저장된 제외목록"만 이번 생성에 적용 (저장 전 편집값은 적용되지 않음)
    desired_editor = [p for p in saved_in_options if p in product_options]

    if "sticker_exclude_products_editor" not in st.session_state:
        st.session_state["sticker_exclude_products_editor"] = desired_editor
    else:
        # 업로드 파일이 바뀌어 옵션 목록이 달라져도 오류가 나지 않게, 현재 옵션에 없는 값은 제거
        st.session_state["sticker_exclude_products_editor"] = [
            p for p in (st.session_state.get("sticker_exclude_products_editor") or [])
            if p in product_options
        ]

    if "sticker_exclude_products_extra" not in st.session_state:
        st.session_state["sticker_exclude_products_extra"] = ",".join(saved_outside)

    with st.expander("🚫 스티커로 출력하지 않을 상품 설정", expanded=False):
        st.caption("선택한 상품은 스티커용지 PDF 생성에서 제외됩니다. (저장하면 다음 실행/다른 파일에도 동일 적용)")

        st.multiselect(
            "제외할 상품 (현재 업로드한 파일에 존재하는 상품)",
            options=product_options,
            key="sticker_exclude_products_editor",
        )

        st.text_input(
            "추가 제외 (옵션에 없는 상품 · 쉼표로 여러개 입력 · 정확히 일치)",
            key="sticker_exclude_products_extra",
            placeholder="예: 고수,딜",
        )

        if st.button("💾 제외목록 저장", use_container_width=True):
            _selected = st.session_state.get("sticker_exclude_products_editor", []) or []
            _extra_text = st.session_state.get("sticker_exclude_products_extra", "") or ""
            _extra = [normalize_text(x) for x in _extra_text.split(",") if normalize_text(x)]

            _merged = []
            _seen = set()
            for _p in (_selected + _extra):
                if _p and _p not in _seen:
                    _merged.append(_p)
                    _seen.add(_p)

            save_sticker_exclude(_merged)
            st.session_state["sticker_exclude_products"] = _merged
            st.success("저장되었습니다. 다음 실행에도 그대로 적용됩니다.")

        st.write(
            "현재 저장된 값:",
            (", ".join(st.session_state.get("sticker_exclude_products", []) or []) or "없음"),
        )

    exclude_set = set(st.session_state.get("sticker_exclude_products", []) or [])

    excluded_stickers = 0

    label_rows = []
    for _, r in summary.iterrows():
        name = str(r["제품명"]).strip()
        qty = _as_int_qty(r["수량"])

        # 제외 상품은 스티커 생성에서 제외
        if name in exclude_set:
            if qty > 0:
                excluded_stickers += qty
            continue

        var = str(r["구분"]).strip()
        label = name if var in ("", "-", "nan", "None") else f"{name}{var}"
        if qty > 0:
            label_rows.append((label, qty))
    label_rows.sort(key=lambda x: x[0])

    sticker_texts: List[str] = []
    for label, qty in label_rows:
        sticker_texts.extend([label] * qty)
    st.caption(f"총 {len(sticker_texts)}개 · 페이지당 65칸 · 글자 {STICKER_FONT_SIZE}pt · 용지 21×29.5cm · 여백 L/R0.4 T1.1 B1.0cm · 라벨 3.82×2.11cm · 가로간격 0.3cm(자동보정) (제외 {excluded_stickers}개)")
    st.download_button(
        "⬇️ 스티커용지 PDF 다운로드",
        data=build_sticker_pdf(sticker_texts),
        file_name="스티커용지.pdf",
        mime="application/pdf",
        use_container_width=True,
    )

    # 수취인별 출력
    st.markdown("---")
    st.subheader("📄 수취인별 출력 ( 새벽 / 익일 )")

    base2 = base.copy()
    base2["배송구분"] = base2["옵션정보"].apply(classify_delivery)
    key_cols = ["구매자명", "수취인명", "통합배송지"]

    grp_deliv = (
        base2.groupby(key_cols)["배송구분"]
        .agg(lambda x: set(x))
        .apply(decide_group_delivery)
        .reset_index()
        .rename(columns={"배송구분": "그룹배송구분"})
    )
    base2 = base2.merge(grp_deliv, on=key_cols, how="left")

    def build_items_for_group(g: pd.DataFrame) -> Tuple[str, str]:
        g = g.sort_index()
        od = OrderedDict()
        for _, r in g.iterrows():
            prod = str(r["제품명"]).strip()
            var = str(r["구분"] or "").strip()
            qty = r["수량"]
            sr = _safe_int(r.get("합산규칙", None))
            if not prod:
                continue
            if var == "":
                var = "-"
            key = (prod, var, sr)
            od[key] = od.get(key, 0.0) + float(qty)

        rows = [{"제품명": p, "구분": v, "수량": q, "합산규칙": sr} for (p, v, sr), q in od.items()]
        rows_df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["제품명", "구분", "수량", "합산규칙"])

        rows_ex = explode_sum_rule_rows(
            rows_df[["제품명", "구분", "수량", "합산규칙"]],
            bundle_units=bundle_units,
            default_unit=default_unit,
        ) if len(rows_df) else rows_df

        od2 = OrderedDict()
        for _, rr in rows_ex.iterrows():
            k2 = (str(rr["제품명"]), str(rr["구분"]))
            od2[k2] = od2.get(k2, 0.0) + float(rr["수량"])

        parts = [f"{pname}/{v} {fmt_qty(q2)}" for (pname, v), q2 in od2.items()]
        recv_name = str(g["수취인명"].iloc[0]).strip()
        return recv_name, ", ".join(parts)

    group_entries = []
    for _, g in base2.groupby(key_cols, sort=False):
        recv_name, items_line = build_items_for_group(g)
        group_entries.append(
            {"그룹배송구분": str(g["그룹배송구분"].iloc[0]), "수취인명": recv_name, "items_line": items_line}
        )

    dawn_entries = [e for e in group_entries if e["그룹배송구분"] == "새벽배송"]
    next_entries = [e for e in group_entries if e["그룹배송구분"] == "익일배송"]

    c1, c2 = st.columns(2)
    with c1:
        st.write(f"새벽배송: {len(dawn_entries)}명")
        st.download_button(
            "⬇️ 새벽배송 수취인별 PDF",
            data=build_recipient_pdf(dawn_entries, footer_prefix="새벽"),
            file_name="새벽배송.pdf",
            mime="application/pdf",
            use_container_width=True,
        )

    with c2:
        st.write(f"익일배송: {len(next_entries)}명")
        st.download_button(
            "⬇️ 익일배송 수취인별 PDF",
            data=build_recipient_pdf(next_entries, footer_prefix="익일"),
            file_name="익일배송.pdf",
            mime="application/pdf",
            use_container_width=True,
        )

    # TC 주문 등록
    st.markdown("---")
    st.subheader("🧾 TC주문_등록양식 ( 새벽 / 익일 )")

    if not TC_TEMPLATE_DEFAULT_PATH.exists():
        st.error("앱 폴더에 'TC주문_등록양식.xlsx' 파일이 없습니다. GitHub에 app.py와 같이 올려주세요.")
    else:
        template_bytes = TC_TEMPLATE_DEFAULT_PATH.read_bytes()

        # 수취인별 출력 순서(원본 등장 순서)
        order_keys_df = base2[key_cols].drop_duplicates(keep="first").copy()

        def _first_nonempty(series: pd.Series) -> str:
            for v in series.tolist():
                if v is None:
                    continue
                s = str(v).strip()
                if s and s.lower() != "nan":
                    return s
            return ""

        grp_info_agg = (
            base2.groupby(key_cols, as_index=False)
            .agg(
                그룹배송구분=("그룹배송구분", "first"),
                수취인연락처=("수취인연락처", _first_nonempty),
                배송메세지=("배송메세지", _first_nonempty),
                구매자명=("구매자명", "first"),
                수취인명=("수취인명", "first"),
                통합배송지=("통합배송지", "first"),
            )
        )
        grp_info = order_keys_df.merge(grp_info_agg, on=key_cols, how="left")

        def make_tc_rows(df: pd.DataFrame, ship: str) -> List[Dict[str, str]]:
            out = []
            ship_type = st.session_state.tc_type_dawn if ship == "새벽배송" else st.session_state.tc_type_next
            ship_type = (ship_type or "").strip() or (TC_TYPE_DAWN_DEFAULT if ship == "새벽배송" else TC_TYPE_NEXT_DEFAULT)

            for _, r in df.iterrows():
                out.append(
                    {
                        "배송요청일": req_day_str,
                        "주문자": str(r["구매자명"] or "").strip(),
                        "수령자": str(r["수취인명"] or "").strip(),
                        "수령자도로명주소": str(r["통합배송지"] or "").strip(),
                        "수령자연락처": str(r.get("수취인연락처", "") or "").strip(),
                        "출입방법": _clean_access_message(r.get("배송메세지", "")),
                        "상품명": TC_PRODUCT_NAME_FIXED,
                        "배송유형": ship_type,
                    }
                )
            return out

        dawn_df = grp_info[grp_info["그룹배송구분"] == "새벽배송"].copy()
        next_df = grp_info[grp_info["그룹배송구분"] == "익일배송"].copy()

        cols = st.columns(2)
        with cols[0]:
            st.write(f"새벽배송 행: {len(dawn_df)} (배송유형: {st.session_state.tc_type_dawn})")
            if len(dawn_df):
                out_bytes = build_tc_excel_bytes(template_bytes, make_tc_rows(dawn_df, "새벽배송"))
                st.download_button(
                    "⬇️ TC주문_등록양식(새벽배송) 엑셀 다운로드",
                    data=out_bytes,
                    file_name="새벽배송_송장.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        with cols[1]:
            st.write(f"익일배송 행: {len(next_df)} (배송유형: {st.session_state.tc_type_next})")
            if len(next_df):
                out_bytes = build_tc_excel_bytes(template_bytes, make_tc_rows(next_df, "익일배송"))
                st.download_button(
                    "⬇️ TC주문_등록양식(익일배송) 엑셀 다운로드",
                    data=out_bytes,
                    file_name="익일배송_송장.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )


def render_product_totals_page():
    st.title("🧾 제품별 합계 (PACK/BOX/EA 규칙 적용)")
    st.caption("PDF 업로드 없이, 엑셀 결과(제품별 개수)를 기반으로 자동 계산합니다.")

    summary_df = st.session_state.get("excel_summary_df")
    default_unit = st.session_state.get("excel_default_unit", "개")

    if summary_df is None or len(summary_df) == 0:
        st.info("먼저 [📥 엑셀 업로드] 페이지에서 엑셀을 업로드해 주세요.")
        if st.button("📥 엑셀 업로드로 이동", use_container_width=True):
            st.session_state["page"] = "excel_results"
            st.rerun()
        return

    # (1) PACK/BOX/EA 규칙 사이드바 (1번 코드 유지)
    if "rules_text" not in st.session_state:
        st.session_state["rules_text"] = load_rules_text()

    allow_decimal_pack = False
    allow_decimal_box = True

    with st.sidebar:
        st.subheader("⚙️ 제품별 합계 표현 규칙")

        with st.expander("🧩 PACK/BOX/EA 규칙", expanded=False):
            up = st.file_uploader("rules.txt 업로드(선택)", type=["txt"], key="rules_uploader")
            if up is not None:
                st.session_state["rules_text"] = up.getvalue().decode("utf-8", errors="ignore")

            st.text_area("규칙", key="rules_text", height=260)

            colA, colB = st.columns(2)
            allow_decimal_pack = colA.checkbox("팩 소수 허용", value=False, key="allow_decimal_pack")
            allow_decimal_box = colB.checkbox("박스 소수 허용", value=True, key="allow_decimal_box")

            with st.form("add_rule_form", clear_on_submit=False):
                st.markdown("**규칙 추가/업데이트**")
                r_type = st.selectbox("TYPE", ["팩", "개", "박스"])
                r_name = st.text_input("상품명(원본 제품명과 동일)", value="")
                r_val = st.text_input("값(PACK=1팩 g, BOX=1박스 kg, EA=1개 g)", value="")
                submitted = st.form_submit_button("추가/업데이트")
                if submitted:
                    st.session_state["rules_text"] = upsert_rule(
                        st.session_state["rules_text"], r_type, r_name, r_val
                    )
                    st.success("규칙 반영 완료!")

            col1, col2 = st.columns(2)
            if col1.button("rules.txt로 저장(로컬용)", key="save_rules_txt"):
                try:
                    save_rules_text(st.session_state["rules_text"])
                    st.success("rules.txt 저장 완료!")
                except Exception as e:
                    st.error(f"저장 실패: {e}")

            col2.download_button(
                "rules.txt 다운로드",
                data=st.session_state["rules_text"].encode("utf-8"),
                file_name="rules.txt",
                mime="text/plain",
            )

    pack_rules, box_rules, ea_rules = parse_rules(st.session_state["rules_text"])

    # (2) 엑셀 요약 기반으로 제품별 합계 계산
    df_long, df_wide = compute_product_totals_from_summary(
        summary_df=summary_df,
        pack_rules=pack_rules,
        box_rules=box_rules,
        ea_rules=ea_rules,
        allow_decimal_pack=allow_decimal_pack,
        allow_decimal_box=allow_decimal_box,
        default_unit=default_unit,
    )

    st.session_state["last_sum_df_long"] = df_long.copy()

    st.subheader("🧾 제품별 합계")
    st.dataframe(df_wide, use_container_width=True, hide_index=True)

    # (3) 다운로드 + 재고등록(1번 코드 그대로)
    try:
        pdf_bytes = make_pdf_bytes(df_wide, "제품별 합계")

        sum_imgs = render_pdf_pages_to_images(pdf_bytes, zoom=3.0)
        sum_png_one = merge_png_pages_to_one(sum_imgs)

        c1, c2, c3 = st.columns(3)
        with c1:
            st.download_button(
                "📄 PDF 다운로드(제품별합계)",
                data=pdf_bytes,
                file_name="제품별합계.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        with c2:
            st.download_button(
                "🖼️ 스크린샷(PNG) 다운로드",
                data=sum_png_one,
                file_name="제품별합계(스크린샷).png",
                mime="image/png",
                use_container_width=True,
            )
        with c3:
            if st.button("📦 재고등록", use_container_width=True):
                st.session_state["show_register_panel"] = True

        if st.session_state.get("show_register_panel"):
            st.markdown("#### 📝 재고등록 (1차/2차/3차)")
            target = st.radio("등록할 차수", ["1차", "2차", "3차"], horizontal=True, key="register_target")
            add_mode = st.checkbox("기존 값에 누적(더하기)", value=False, key="register_add_mode")

            colR1, colR2 = st.columns([1, 3])
            with colR1:
                do_reg = st.button("✅ 등록", use_container_width=True, key="do_register_btn")
            with colR2:
                st.caption("※ 재고관리 표에 **이미 존재하는 상품명만** 등록됩니다. (없는 상품은 제외)")

            if do_reg:
                sum_df = st.session_state.get("last_sum_df_long")
                updated, skipped = register_sum_to_inventory(sum_df, target_col=target, add_mode=add_mode)
                sync_ok, sync_msg = sync_inventory_to_google_sheet(sum_df, target_col=target, add_mode=add_mode)
                st.session_state["show_register_panel"] = False

                if skipped:
                    st.warning("등록 제외(재고관리 상품명 없음): " + ", ".join(sorted(set(skipped))))
                st.success(f"{target}에 등록 완료! (반영 행: {updated})")
                if sync_ok:
                    st.info(sync_msg)
                else:
                    st.warning(sync_msg)
                st.info("📦 사이드바의 '재고관리'로 이동하면 확인할 수 있어요.")

        if Image is None and len(sum_imgs) > 1:
            st.warning("⚠️ Pillow(PIL)가 없어 제품별합계 스크린샷은 1페이지만 PNG로 저장됩니다. 전체를 1장으로 합치려면 Pillow 설치가 필요합니다.")
    except Exception as e:
        st.error(f"제품별 합계 PDF/PNG 생성 실패: {e} (fonts/NanumGothic.ttf 또는 pymupdf 확인)")


def render_inventory_page():
    st.title("재고관리")

    # ---- 📁 내보내기 폴더(재고관리에서만 표시) ----
    with st.sidebar:
        with st.expander("📁 내보내기 폴더", expanded=False):
            dates = list_export_dates()
            if not dates:
                st.caption("내보내기 기록이 없습니다.")
            else:
                last = st.session_state.get("last_export_date")
                if last:
                    st.caption(f"마지막 내보내기: {last}")

                st.caption("※ 삭제하면 복구할 수 없습니다.")
                for d in dates:
                    data = read_export_xlsx_bytes(d)
                    row1, row2 = st.columns([3, 1])

                    with row1:
                        if data is None:
                            st.caption(f"📁 {d} (파일 없음)")
                        else:
                            st.download_button(
                                label=f"⬇️ {d} 재고표(.xlsx)",
                                data=data,
                                file_name=f"재고표_{d}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key=f"export_dl_{d}",
                            )

                    with row2:
                        if st.button("🗑️", use_container_width=True, key=f"export_del_{d}"):
                            ok = False
                            try:
                                ok = delete_export_date(d)
                            except Exception:
                                ok = False

                            if ok:
                                if st.session_state.get("last_export_date") == d:
                                    st.session_state["last_export_date"] = None
                                st.session_state["inventory_toast"] = f"{d} 내보내기 삭제 완료!"
                                st.rerun()
                            else:
                                st.error("삭제 실패: 폴더/파일을 확인해주세요.")

    msg = st.session_state.pop("inventory_toast", None)
    if msg:
        st.success(msg)

    if "inventory_df" not in st.session_state:
        st.session_state["inventory_df"] = load_inventory_df()
    if "inventory_editor_version" not in st.session_state:
        st.session_state["inventory_editor_version"] = 0

    df_view = compute_inventory_df(st.session_state["inventory_df"])
    df_view = sort_inventory_df(df_view).reset_index(drop=True)
    df_view = df_view[df_view["상품명"].astype(str).str.strip() != ""].reset_index(drop=True)

    # -------------------- 스타일 (1번 코드 유지) --------------------
    def _remain_bg(v):
        try:
            x = float(v)
        except Exception:
            return ""
        if x < 0:
            return "background-color: #ffb3b3;"  # 연한 빨강
        if 0 <= x <= 10:
            return "background-color: #ffd6e7;"  # 연분홍
        if x >= 30:
            return "background-color: #d6ecff;"  # 연파랑
        return ""

    st.markdown(
        """
        <style>
        /* 헤더 텍스트 Bold */
        div[data-testid="stDataEditor"] .ag-header-cell[col-id="상품명"] .ag-header-cell-text,
        div[data-testid="stDataEditor"] .ag-header-cell[col-id="보유수량"] .ag-header-cell-text,
        div[data-testid="stDataEditor"] .ag-header-cell[col-id="남은수량"] .ag-header-cell-text,
        div[data-testid="stDataFrame"]  .ag-header-cell[col-id="상품명"] .ag-header-cell-text,
        div[data-testid="stDataFrame"]  .ag-header-cell[col-id="보유수량"] .ag-header-cell-text,
        div[data-testid="stDataFrame"]  .ag-header-cell[col-id="남은수량"] .ag-header-cell-text {
            font-weight: 800 !important;
        }

        /* 셀 값 Bold(폴백) */
        div[data-testid="stDataEditor"] .ag-cell[col-id="상품명"],
        div[data-testid="stDataEditor"] .ag-cell[col-id="보유수량"],
        div[data-testid="stDataEditor"] .ag-cell[col-id="남은수량"],
        div[data-testid="stDataEditor"] .ag-cell[col-id="상품명"] .ag-cell-value,
        div[data-testid="stDataEditor"] .ag-cell[col-id="보유수량"] .ag-cell-value,
        div[data-testid="stDataEditor"] .ag-cell[col-id="남은수량"] .ag-cell-value,
        div[data-testid="stDataFrame"]  .ag-cell[col-id="상품명"],
        div[data-testid="stDataFrame"]  .ag-cell[col-id="보유수량"],
        div[data-testid="stDataFrame"]  .ag-cell[col-id="남은수량"],
        div[data-testid="stDataFrame"]  .ag-cell[col-id="상품명"] .ag-cell-value,
        div[data-testid="stDataFrame"]  .ag-cell[col-id="보유수량"] .ag-cell-value,
        div[data-testid="stDataFrame"]  .ag-cell[col-id="남은수량"] .ag-cell-value {
            font-weight: 800 !important;
        }

        /* ✅ 재고표 데이터(셀) 전체 왼쪽 정렬 (숫자 포함) */
        div[data-testid="stDataEditor"] .ag-center-cols-container .ag-cell[col-id],
        div[data-testid="stDataFrame"]  .ag-center-cols-container .ag-cell[col-id] {
            text-align: left !important;
            justify-content: flex-start !important;
        }
        div[data-testid="stDataEditor"] .ag-center-cols-container .ag-cell[col-id] .ag-cell-wrapper,
        div[data-testid="stDataFrame"]  .ag-center-cols-container .ag-cell[col-id] .ag-cell-wrapper {
            justify-content: flex-start !important;
            width: 100% !important;
        }
        div[data-testid="stDataEditor"] .ag-center-cols-container .ag-cell[col-id] .ag-cell-value,
        div[data-testid="stDataFrame"]  .ag-center-cols-container .ag-cell[col-id] .ag-cell-value {
            text-align: left !important;
            width: 100% !important;
        }
        div[data-testid="stDataEditor"] .ag-center-cols-container .ag-cell[col-id] input,
        div[data-testid="stDataFrame"]  .ag-center-cols-container .ag-cell[col-id] input {
            text-align: left !important;
        }

        /* 숫자 기본 오른쪽 정렬 클래스 강제 override */
        div[data-testid="stDataEditor"] .ag-cell.ag-right-aligned,
        div[data-testid="stDataFrame"]  .ag-cell.ag-right-aligned,
        div[data-testid="stDataEditor"] .ag-cell.ag-number-cell,
        div[data-testid="stDataFrame"]  .ag-cell.ag-number-cell {
            text-align: left !important;
        }
        div[data-testid="stDataEditor"] .ag-cell.ag-right-aligned .ag-cell-wrapper,
        div[data-testid="stDataFrame"]  .ag-cell.ag-right-aligned .ag-cell-wrapper,
        div[data-testid="stDataEditor"] .ag-cell.ag-number-cell .ag-cell-wrapper,
        div[data-testid="stDataFrame"]  .ag-cell.ag-number-cell .ag-cell-wrapper {
            justify-content: flex-start !important;
            width: 100% !important;
        }

        /* (선택) 헤더도 왼쪽 정렬 */
        div[data-testid="stDataEditor"] .ag-header-cell .ag-header-cell-label,
        div[data-testid="stDataFrame"]  .ag-header-cell .ag-header-cell-label {
            justify-content: flex-start !important;
        }
        div[data-testid="stDataEditor"] .ag-header-cell-text,
        div[data-testid="stDataFrame"]  .ag-header-cell-text {
            text-align: left !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    df_display = df_view.copy()

    def _fmt_num(v):
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return "0"
        try:
            x = float(v)
            if abs(x) < 1e-12:
                x = 0.0
            if float(x).is_integer():
                return str(int(round(x)))
            return format(x, "g")
        except Exception:
            s = str(v).strip()
            return s if s else "0"

    for c in ["재고", "입고", "보유수량", "1차", "2차", "3차", "주문수량", "남은수량"]:
        if c in df_display.columns:
            df_display[c] = df_display[c].map(_fmt_num)

    def _remain_bg_any(v):
        try:
            x = float(str(v).replace(",", "").strip())
        except Exception:
            return ""
        return _remain_bg(x)

    _styler = df_display.style
    if hasattr(_styler, "map"):
        _styler = _styler.map(_remain_bg_any, subset=["남은수량"])
    else:
        _styler = _styler.applymap(_remain_bg_any, subset=["남은수량"])

    df_styler = _styler.set_properties(
        subset=["상품명", "보유수량", "남은수량"],
        **{"font-weight": "800"}
    )

    st.markdown("### 재고표 (수정/추가/삭제 가능)")

    ver = int(st.session_state.get("inventory_editor_version", 0))
    editor_key = f"inventory_editor_{ver}"

    edited_raw = st.data_editor(
        df_styler,
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        disabled=["보유수량", "주문수량", "남은수량"],
        column_config={
            "상품명": st.column_config.TextColumn("상품명", required=True),
            "재고": st.column_config.TextColumn("재고"),
            "입고": st.column_config.TextColumn("입고"),
            "보유수량": st.column_config.TextColumn("보유수량"),
            "1차": st.column_config.TextColumn("1차"),
            "2차": st.column_config.TextColumn("2차"),
            "3차": st.column_config.TextColumn("3차"),
            "주문수량": st.column_config.TextColumn("주문수량"),
            "남은수량": st.column_config.TextColumn("남은수량"),
        },
        key=editor_key,
    )
    edited_raw = edited_raw.copy() if isinstance(edited_raw, pd.DataFrame) else pd.DataFrame(edited_raw)

    def _base_view(df: pd.DataFrame) -> pd.DataFrame:
        base_cols = ["상품명", "재고", "입고", "1차", "2차", "3차"]
        dd = df.copy()
        for c in base_cols:
            if c not in dd.columns:
                dd[c] = "" if c == "상품명" else 0
        dd["상품명"] = dd["상품명"].fillna("").astype(str).str.strip()
        for c in ["재고", "입고", "1차", "2차", "3차"]:
            dd[c] = pd.to_numeric(dd[c], errors="coerce").fillna(0.0)
        return dd[base_cols].reset_index(drop=True)

    df_base_new = _base_view(edited_raw)
    df_base_new = df_base_new[df_base_new["상품명"].astype(str).str.strip() != ""].reset_index(drop=True)

    dup = df_base_new["상품명"][df_base_new["상품명"].duplicated(keep=False)]
    if len(dup) > 0:
        st.warning(f"⚠️ 상품명이 중복된 행이 있습니다: {', '.join(sorted(set(dup.astype(str))))}")

    colA, colB, colC = st.columns([1, 1, 1])

    if colA.button("💾 저장", use_container_width=True):
        df_save = compute_inventory_df(df_base_new)
        df_save = sort_inventory_df(df_save).reset_index(drop=True)
        df_save = df_save[df_save["상품명"].astype(str).str.strip() != ""].reset_index(drop=True)

        st.session_state["inventory_df"] = df_save
        save_inventory_df(df_save)

        st.session_state["inventory_editor_version"] = ver + 1
        st.session_state["inventory_toast"] = "저장 완료!"
        st.rerun()

    if colB.button("↻ 초기화(0으로)", use_container_width=True):
        base = zero_numeric_inventory_fields(df_base_new)
        base = compute_inventory_df(base)
        base = sort_inventory_df(base).reset_index(drop=True)

        st.session_state["inventory_df"] = base
        save_inventory_df(base)

        st.session_state["inventory_editor_version"] = ver + 1
        st.session_state["inventory_toast"] = "현재 상품명은 유지하고 숫자만 0으로 초기화했습니다!"
        st.rerun()

    if colC.button("📤 내보내기", use_container_width=True):
        df_export = compute_inventory_df(df_base_new)
        df_export = sort_inventory_df(df_export).reset_index(drop=True)
        df_export = df_export[df_export["상품명"].astype(str).str.strip() != ""].reset_index(drop=True)

        try:
            date_str, _ = export_inventory_snapshot(df_export)

            df_roll = df_export.copy()
            remain = pd.to_numeric(df_roll["남은수량"], errors="coerce").fillna(0.0)
            df_roll["재고"] = remain.clip(lower=0.0)  # ✅ 음수는 재고로 이관하지 않음
            for c in ["입고", "1차", "2차", "3차"]:
                df_roll[c] = 0.0

            df_roll = df_roll[["상품명", "재고", "입고", "1차", "2차", "3차"]]
            df_roll = compute_inventory_df(df_roll)
            df_roll = sort_inventory_df(df_roll).reset_index(drop=True)
            df_roll = df_roll[df_roll["상품명"].astype(str).str.strip() != ""].reset_index(drop=True)

            st.session_state["inventory_df"] = df_roll
            save_inventory_df(df_roll)

            st.session_state["inventory_editor_version"] = ver + 1
            st.session_state["inventory_toast"] = (
                f"내보내기 완료! 남은수량을 재고로 이관(음수는 0 처리)했고, 나머지는 0으로 초기화했습니다. "
                f"(사이드바 ▶ 📁 내보내기 폴더 ▶ {date_str})"
            )
            st.session_state["last_export_date"] = date_str
            st.rerun()
        except Exception as e:
            st.error(f"내보내기 실패: {e}")



# =====================================================
# Sales Calc Page (매출계산)  ✅ 2.py 기능 통합
# =====================================================


# =====================================================
# Extracted page modules are imported above
# =====================================================

# =====================================================
page = st.session_state.get("page", "excel_results")
if page == "mapping_rules":
    render_mapping_rules_page()
elif page == "product_totals":
    render_product_totals_page()
elif page == "inventory":
    render_inventory_page()
elif page == "invoice_register":
    render_invoice_register_page()
elif page == "bulk_stock":
    render_bulk_stock_page()
elif page == "sales_calc":
    render_sales_calc_page()
else:
    render_excel_results_page()

from typing import Optional
import requests
from config import APP_DATA_DIR, now_prefix_kst

def fmt_qty_no_zero(x):
    try:
        x = float(x)
        if abs(x) < 1e-9:
            return "0"
        return str(int(x)) if x.is_integer() else str(x)
    except Exception:
        return str(x)


def render_bulk_stock_page():
    import io
    import json
    import os
    import hashlib
    from dataclasses import dataclass
    from datetime import datetime
    import time
    from decimal import Decimal, InvalidOperation, getcontext
    import re
    from pathlib import Path
    from typing import Dict, Any, Tuple, List

    import pandas as pd
    import streamlit as st
    from openpyxl import load_workbook

    # ============================
    # Persistent config (best effort)
    # ============================
    APP_DIR = Path(__file__).parent

    # Prefer a writable data dir (avoid writing into repo root directly)
    DATA_DIR = Path(os.environ.get("STOCKAPP_DATA_DIR") or str(Path(globals().get("APP_DATA_DIR", APP_DIR)) / "stock_bulk"))
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    CONFIG_FILE = DATA_DIR / "stock_config.json"
    CONFIG_BAK = DATA_DIR / "stock_config.bak.json"

    EXCEL_CONFIG_SHEET = "_STOCKAPP_CONFIG"
    EXCEL_CONFIG_CELL = "A1"

    DEFAULT_PRODUCTS = [
        {"name": "엔다이브", "keyword": "엔다이브"},
        {"name": "샬롯", "keyword": "샬롯"},
        {"name": "아스파라", "keyword": "생 아스파라"},
        {"name": "화이트아스파라", "keyword": "화이트 아스파라"},
        {"name": "미니양배추", "keyword": "미니양배추"},
        {"name": "양송이", "keyword": "양송이"},
        {"name": "새송이", "keyword": "새송이"},
        {"name": "느타리", "keyword": "느타리"},
        {"name": "팽이", "keyword": "팽이"},
        {"name": "흙당근", "keyword": "흙당근"},
        {"name": "브로콜리", "keyword": "브로콜리"},
        {"name": "컬리플라워", "keyword": "컬리플라워"},
        {"name": "줄기샐러리", "keyword": "줄기샐러리"},
        {"name": "오렌지", "keyword": "오렌지"},
        {"name": "자몽", "keyword": "자몽"},
        {"name": "레몬", "keyword": "레몬"},
        {"name": "라임", "keyword": "라임"},
        {"name": "양상추", "keyword": "양상추"},
        {"name": "알배기", "keyword": "알배기"},
        {"name": "방울토마토", "keyword": "방울토마토"},
        {"name": "완숙토마토", "keyword": "완숙토마토"},
        {"name": "아보카도", "keyword": "아보카도"},
        {"name": "식용꽃", "keyword": "식용꽃"},
        {"name": "청피망", "keyword": "청피망"},
        {"name": "미니파프리카", "keyword": "미니 파프리카"},
        {"name": "삼색파프리카", "keyword": "삼색 파프리카"},
        {"name": "비트", "keyword": "비트"},
        {"name": "콜라비", "keyword": "콜라비"},
        {"name": "파세리", "keyword": "파세리"},
        {"name": "깐마늘", "keyword": "깐마늘"},
        {"name": "단호박", "keyword": "단호박"},
        {"name": "쥬키니", "keyword": "쥬키니"},
        {"name": "가지", "keyword": "가지"},
        {"name": "백오이", "keyword": "백오이"},
    ]


    def _default_config() -> Dict[str, Any]:
        return {
            "version": 9,
            "inventory_column": "재고수량",
            "name_column": "상품명",
            "products": [],
            "rules": {},  # {base_product: [ {keyword, mode, value, table}, ... ]}
            "ref_qty": {},  # {product_name: "참고수량"}
            "recognition_logic": DEFAULT_RECOGNITION_LOGIC.copy(),
        }


    def _atomic_write(path: Path, text: str, encoding: str = "utf-8") -> None:
        tmp = path.with_suffix(path.suffix + ".tmp")
        tmp.write_text(text, encoding=encoding)
        tmp.replace(path)


    def load_config() -> Dict[str, Any]:
        for p in [CONFIG_FILE, CONFIG_BAK]:
            if p.exists():
                try:
                    cfg = json.loads(p.read_text(encoding="utf-8"))
                    if not isinstance(cfg, dict):
                        continue
                    cfg.setdefault("version", 9)
                    cfg.setdefault("inventory_column", "재고수량")
                    cfg.setdefault("name_column", "상품명")
                    cfg.setdefault("products", [])
                    cfg.setdefault("rules", {})
                    cfg.setdefault("ref_qty", {})
                    cfg.setdefault("recognition_logic", DEFAULT_RECOGNITION_LOGIC.copy())
                    return cfg
                except Exception:
                    continue
        return _default_config()


    def save_config(cfg: Dict[str, Any]) -> None:
        cfg["version"] = 9
        txt = json.dumps(cfg, ensure_ascii=False, indent=2)
        if CONFIG_FILE.exists():
            try:
                _atomic_write(CONFIG_BAK, CONFIG_FILE.read_text(encoding="utf-8"))
            except Exception:
                pass
        _atomic_write(CONFIG_FILE, txt)


    # ----------------------------
    # Excel helpers
    # ----------------------------
    def find_header_row_and_columns(ws, name_col: str, inv_col: str, max_scan_rows: int = 15) -> Tuple[int, int, int]:
        for r in range(1, max_scan_rows + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            row_str = [str(v).strip() if v is not None else "" for v in row_vals]
            if name_col in row_str and inv_col in row_str:
                name_idx = row_str.index(name_col) + 1
                inv_idx = row_str.index(inv_col) + 1
                return r, name_idx, inv_idx
        raise ValueError(f"헤더에서 '{name_col}' 또는 '{inv_col}' 컬럼을 찾지 못했습니다.")


    def to_number(x) -> float:
        if x is None:
            return 0.0
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip()
        if s == "":
            return 0.0
        s = s.replace(",", "")
        try:
            return float(s)
        except Exception:
            return 0.0


    def parse_input_number(x) -> float:
        """Parse user input from data_editor (string or number)."""
        if x is None:
            return 0.0
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip()
        if s == "":
            return 0.0
        s = s.replace(",", "")
        try:
            return float(s)
        except Exception:
            return 0.0


    def qty_key(q: float) -> str:
        """Normalize input quantity to mapping key string."""
        try:
            qf = float(q)
        except Exception:
            return str(q)
        if abs(qf - int(qf)) < 1e-9:
            return str(int(qf))
        return str(qf)


    def parse_map_string(s: str) -> Dict[str, int]:
        """
        Accept formats:
          - "2=5, 3=7"
          - "2:5\n3:7"
          - "2 -> 5"
        Return {"2":5, "3":7}
        Values are stored as INT to avoid 5.0 display.
        """
        if not s:
            return {}
        txt = str(s).strip()
        if not txt:
            return {}
        txt = txt.replace("\n", ",")
        txt = txt.replace("→", "=").replace("->", "=").replace(":", "=")
        out: Dict[str, int] = {}
        for chunk in [x.strip() for x in txt.split(",") if x.strip()]:
            if "=" not in chunk:
                continue
            k, v = chunk.split("=", 1)
            k = k.strip()
            v = v.strip()
            if not k or not v:
                continue

            # allow wildcard key "*" (입력값 무관 고정 적용)
            if k == "*":
                try:
                    out["*"] = int(round(float(v)))
                except Exception:
                    pass
                continue
            try:
                kf = float(k)
                kk = str(int(kf)) if abs(kf - int(kf)) < 1e-9 else str(kf)
                out[kk] = int(round(float(v)))
            except Exception:
                continue
        return out


    def fmt_int(v) -> str:
        try:
            return str(int(round(float(v))))
        except Exception:
            return str(v)


    def fmt_qty_for_memo(v: float) -> str:
        """Pretty formatting for memo txt (avoid trailing .0)."""
        try:
            vf = float(v)
        except Exception:
            return str(v)
        if abs(vf - int(vf)) < 1e-9:
            return str(int(vf))
        s = f"{vf:.12f}".rstrip("0").rstrip(".")
        return s if s else "0"


    # ----------------------------
    # Inventory total (display-only)
    # ----------------------------
    getcontext().prec = 28

    _RE_KG = re.compile(r"(\d+(?:\.\d+)?)\s*kg", re.IGNORECASE)
    _RE_G = re.compile(r"(\d+(?:\.\d+)?)\s*g", re.IGNORECASE)
    _RE_PACK = re.compile(r"(\d+(?:\.\d+)?)\s*팩")

    _RE_PACK_EN = re.compile(r"(\d+(?:\.\d+)?)\s*pack", re.IGNORECASE)
    _RE_BONG = re.compile(r"(\d+(?:\.\d+)?)\s*봉")
    _RE_TONG = re.compile(r"(\d+(?:\.\d+)?)\s*통")
    _RE_EA = re.compile(r"(\d+(?:\.\d+)?)\s*개")
    _RE_BOX = re.compile(r"(\d+(?:\.\d+)?)\s*(박스|box)", re.IGNORECASE)


    # ----------------------------
    # Recognition logic (unit parsing)
    # ----------------------------
    # 각 항목: priority(낮을수록 먼저), output_unit(표시 단위), multiplier(숫자에 곱), aliases(인식할 문자열들)
    DEFAULT_RECOGNITION_LOGIC: List[Dict[str, Any]] = [
        {"priority": 10, "output_unit": "단",   "multiplier": "1",     "aliases": ["단"]},
        {"priority": 20, "output_unit": "팩",   "multiplier": "1",     "aliases": ["팩", "pack"]},
        {"priority": 30, "output_unit": "봉",   "multiplier": "1",     "aliases": ["봉"]},
        {"priority": 40, "output_unit": "통",   "multiplier": "1",     "aliases": ["통"]},
        {"priority": 50, "output_unit": "개",   "multiplier": "1",     "aliases": ["개", "ea"]},
        {"priority": 60, "output_unit": "박스", "multiplier": "1",     "aliases": ["박스", "box"]},
        # g는 kg로 환산(예: 500g -> 0.5kg)
        {"priority": 90, "output_unit": "kg",   "multiplier": "0.001", "aliases": ["g", "그램"]},
        {"priority": 100,"output_unit": "kg",   "multiplier": "1",     "aliases": ["kg", "킬로", "키로"]},
    ]

    def _normalize_recognition_logic(logic: Any) -> List[Dict[str, Any]]:
        """config에 저장된 recognition_logic를 안전하게 정규화합니다."""
        if not isinstance(logic, list):
            logic = []
        cleaned: List[Dict[str, Any]] = []
        for it in logic:
            if not isinstance(it, dict):
                continue
            unit = str(it.get("output_unit") or it.get("unit") or "").strip()
            if not unit:
                continue

            # priority
            try:
                pr = int(it.get("priority", 999))
            except Exception:
                pr = 999

            # multiplier
            mult_raw = it.get("multiplier", "1")
            try:
                mult = str(mult_raw).strip()
                Decimal(mult)  # validate
            except Exception:
                mult = "1"

            aliases = it.get("aliases") or it.get("patterns") or []
            if isinstance(aliases, str):
                aliases = [a.strip() for a in aliases.split(",") if a.strip()]
            if not isinstance(aliases, list):
                aliases = []
            aliases = [str(a).strip() for a in aliases if str(a).strip()]
            if not aliases:
                continue

            cleaned.append({"priority": pr, "output_unit": unit, "multiplier": mult, "aliases": aliases})

        cleaned.sort(key=lambda x: (x.get("priority", 999), len(str(x.get("output_unit","")))))
        return cleaned

    def _get_recognition_logic(cfg: Dict[str, Any]) -> List[Dict[str, Any]]:
        return _normalize_recognition_logic((cfg or {}).get("recognition_logic")) or DEFAULT_RECOGNITION_LOGIC.copy()



    def _cell_to_decimal(v) -> Decimal:
        if v is None:
            return Decimal("0")
        if isinstance(v, (int, float)):
            # Use str() to preserve "0.1" rather than binary float
            return Decimal(str(v))
        s = str(v).strip().replace(",", "")
        if not s:
            return Decimal("0")
        try:
            return Decimal(s)
        except Exception:
            return Decimal("0")


    def _parse_factor_and_unit(name_str: str, recognition_logic: List[Dict[str, Any]] = None) -> Tuple[Decimal, str]:
        """
        상품명(또는 규칙 키워드) 문자열에서 "숫자+단위"를 찾아 (factor, unit)을 반환합니다.

        - recognition_logic 기반으로 처리합니다.
          각 항목: priority, output_unit, multiplier, aliases
            예) {"output_unit":"단","multiplier":"1","aliases":["단"]}

        - 숫자+단위가 있으면 factor = 숫자 * multiplier
        - 단위만 있으면 factor = 1 * multiplier

        반환 unit은 output_unit 입니다. (예: g는 kg로 환산되어 unit='kg')
        """
        s = name_str or ""
        logic = _normalize_recognition_logic(recognition_logic) if recognition_logic is not None else DEFAULT_RECOGNITION_LOGIC.copy()

        for rule in logic:
            unit = str(rule.get("output_unit", "")).strip()
            if not unit:
                continue
            try:
                mult = Decimal(str(rule.get("multiplier", "1")).strip() or "1")
            except Exception:
                mult = Decimal("1")

            aliases = rule.get("aliases") or []
            if not isinstance(aliases, list):
                aliases = [str(aliases)]

            for alias in aliases:
                alias = str(alias).strip()
                if not alias:
                    continue

                # 1) 숫자+단위
                try:
                    if alias.isascii():
                        m = re.search(rf"(\d+(?:\.\d+)?)\s*{re.escape(alias)}\b", s, flags=re.IGNORECASE)
                    else:
                        m = re.search(rf"(\d+(?:\.\d+)?)\s*{re.escape(alias)}", s)
                except Exception:
                    m = None

                if m:
                    try:
                        n = Decimal(m.group(1))
                    except Exception:
                        n = Decimal("1")
                    return (n * mult), unit

                # 2) 단위만 있는 경우
                try:
                    if alias.isascii():
                        if re.search(rf"\b{re.escape(alias)}\b", s, flags=re.IGNORECASE):
                            return (Decimal("1") * mult), unit
                    else:
                        if alias in s:
                            return (Decimal("1") * mult), unit
                except Exception:
                    pass

        return Decimal("1"), ""


    def _fmt_decimal(d: Decimal, max_decimals: int = 3) -> str:
        try:
            q = Decimal("1") if max_decimals <= 0 else (Decimal("1") / (Decimal(10) ** max_decimals))
            d2 = d.quantize(q)  # rounding
        except Exception:
            d2 = d

        s = format(d2, "f")
        s = s.rstrip("0").rstrip(".")
        return s if s else "0"


    def compute_stock_display_map(xlsx_bytes: bytes, cfg: Dict[str, Any]) -> Dict[str, str]:
        """
        Build display stock totals per base product using **규칙관리 키워드**를 참고하여 계산합니다.

        - 각 기준상품(base)에 대해 규칙(키워드)을 이용해 옵션 단위를 파악합니다.
          예) '1kg', '500g', '100g' -> kg(합산),  '5팩' -> 팩,  '6통' -> 통
        - 엑셀의 각 행(상품명)을 규칙 키워드와 매칭하여, 재고수량 * (옵션 단위 수량) 를 합산합니다.
          예) 양상추6통:3개 + 양상추1통:3개 => 3*6 + 3*1 = 21통

        Stock 없는 경우는 '0'으로 표시합니다.
        """
        inv_col = cfg.get("inventory_column", "재고수량")
        name_col = cfg.get("name_column", "상품명")
        recog_logic = _get_recognition_logic(cfg)

        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        ws = wb.active
        header_row, name_idx, inv_idx = find_header_row_and_columns(ws, name_col=name_col, inv_col=inv_col)

        products = cfg.get("products", []) or []
        base_kw: Dict[str, str] = {}
        for p in products:
            bn = str(p.get("name", "")).strip()
            kw = str(p.get("keyword") or p.get("name") or "").strip()
            if bn:
                base_kw[bn] = kw or bn

        bases_sorted = sorted([(bn, base_kw.get(bn, bn)) for bn in base_kw.keys()], key=lambda x: len(str(x[1] or "")), reverse=True)

        rules_map = cfg.get("rules", {}) or {}

        def _scoped_tokens(base_keyword: str, rule_keyword: str) -> List[str]:
            base_keyword = str(base_keyword or "").strip()
            rule_keyword = str(rule_keyword or "").strip()
            if not rule_keyword:
                return []
            if not base_keyword:
                return [rule_keyword]
            # If user already typed full keyword incl base keyword, treat as absolute substring match
            if base_keyword in rule_keyword:
                return [rule_keyword]
            # Otherwise require BOTH base keyword and rule keyword
            return [base_keyword, rule_keyword]

        # Build matchers from rules (per base)
        matchers_by_base: Dict[str, List[Dict[str, Any]]] = {}
        unit_pref: Dict[str, str] = {}
        unit_seen: Dict[str, str] = {}

        for bn, bkw in bases_sorted:
            rs = rules_map.get(bn, []) or []
            ms: List[Dict[str, Any]] = []

            unit_counts: Dict[str, int] = {}
            for i, r in enumerate(rs):
                rr = Rule.from_dict(r)
                if not rr.keyword:
                    continue

                tokens = _scoped_tokens(bkw, rr.keyword)
                if not tokens:
                    continue

                factor, unit = _parse_factor_and_unit(rr.keyword, recog_logic)

                ms.append(
                    {
                        "tokens": tokens,
                        "keyword": rr.keyword,
                        "factor": factor,
                        "unit": unit,
                        "order": i,
                    }
                )

                if unit:
                    unit_counts[unit] = unit_counts.get(unit, 0) + 1

            # Sort by specificity (longer/ more tokens first)
            ms.sort(
                key=lambda m: (sum(len(t) for t in m["tokens"]), len(m["tokens"]), len(m["keyword"])),
                reverse=True,
            )
            matchers_by_base[bn] = ms

            # Preferred unit: kg wins if any kg/g exists; else most frequent unit in rules
            if unit_counts.get("kg", 0) > 0:
                unit_pref[bn] = "kg"
            elif unit_counts:
                unit_pref[bn] = sorted(unit_counts.items(), key=lambda kv: (kv[1], len(kv[0])), reverse=True)[0][0]
            else:
                unit_pref[bn] = ""

        totals: Dict[str, Decimal] = {}

        def _match_tokens(tokens: List[str], name_str: str) -> bool:
            return bool(tokens) and all((t in name_str) for t in tokens)

        for r in range(header_row + 1, ws.max_row + 1):
            name_val = ws.cell(r, name_idx).value
            if name_val is None or str(name_val).strip() == "":
                continue

            name_str = str(name_val)

            # determine base product by base keyword
            base_name = None
            for bn, kw in bases_sorted:
                if kw and kw in name_str:
                    base_name = bn
                    break
            if base_name is None:
                continue

            inv_qty = _cell_to_decimal(ws.cell(r, inv_idx).value)
            if inv_qty == 0:
                continue

            # 1) Try rules-based matching first
            chosen = None
            for m in matchers_by_base.get(base_name, []):
                if _match_tokens(m["tokens"], name_str):
                    chosen = m
                    break

            has_rules = bool(matchers_by_base.get(base_name))
            if chosen and chosen.get("unit"):
                factor = chosen["factor"]
                unit = chosen["unit"]
            else:
                # ✅ 규칙이 있는 기준상품은 "규칙에 매칭되는 옵션"만 합산 (단위 혼합 방지)
                # 규칙이 있는데 어떤 키워드에도 안 걸리면, 해당 행은 재고 합산에서 제외합니다.
                if has_rules and chosen is None:
                    continue
                # 2) Fallback: parse from actual name (kg/g/팩/봉/통/개/박스)
                factor, unit = _parse_factor_and_unit(name_str, recog_logic)

            # Track unit seen (if rules did not define a preferred unit)
            pref = unit_pref.get(base_name, "") or ""
            if pref:
                unit_seen[base_name] = pref
            elif unit:
                unit_seen[base_name] = unit_seen.get(base_name) or unit

            # Sum: inv_qty * factor (factor already normalized: g -> kg)
            totals[base_name] = totals.get(base_name, Decimal("0")) + (inv_qty * factor)

        out: Dict[str, str] = {}
        for bn, _kw in bases_sorted:
            total = totals.get(bn, Decimal("0"))
            if total == 0:
                out[bn] = "0"
                continue

            unit = unit_pref.get(bn, "") or unit_seen.get(bn, "") or ""
            if unit == "kg":
                out[bn] = f"{_fmt_decimal(total, max_decimals=3)}kg"
            else:
                s = _fmt_decimal(total, max_decimals=3)
                out[bn] = f"{s}{unit}" if unit else s

        return out


    # ----------------------------
    # Rule model
    # ----------------------------
    @dataclass
    class Rule:
        keyword: str
        mode: str   # "mul" or "map"
        value: float
        table: Dict[str, int]

        @staticmethod
        def from_dict(d: Dict[str, Any]) -> "Rule":
            keyword = str(d.get("keyword", "")).strip()
            mode_raw = str(d.get("mode", "mul")).strip()

            # legacy fixed -> map wildcard (입력값 무관하게 적용)
            value_raw = float(d.get("value", 1.0) or 0.0)

            table = d.get("table") or {}
            if isinstance(table, str):
                table = parse_map_string(table)
            if not isinstance(table, dict):
                table = {}

            t2: Dict[str, int] = {}
            for k, v in table.items():
                try:
                    t2[str(k)] = int(round(float(v)))
                except Exception:
                    continue

            # migrate fixed -> map wildcard
            if mode_raw == "fixed":
                mode_raw = "map"
                t2.setdefault("*", int(round(value_raw)))

            # ignore legacy round key silently
            if mode_raw not in ("mul", "map"):
                mode_raw = "mul"

            return Rule(keyword=keyword, mode=mode_raw, value=value_raw, table=t2)


    def build_actions(cfg: Dict[str, Any], inputs: Dict[str, float]) -> Tuple[List[Dict[str, Any]], pd.DataFrame]:
        """
        Build merged actions like:
          [{"match_tokens":["엔다이브","1kg"],"display":"엔다이브 & 1kg","delta":5.0,"bases":[...]}...]

        ✅ 규칙 키워드 자동 '기준상품 스코프' 처리
        - 규칙 키워드에 **기준상품 키워드(예: 엔다이브)**가 이미 포함되어 있으면: 그대로(단일 포함문자열) 매칭
          예) "엔다이브1kg" -> '엔다이브1kg'가 포함된 상품명만 매칭
        - 규칙 키워드에 기준상품 키워드가 **없으면**: (기준상품 키워드 AND 규칙 키워드) 둘 다 포함된 상품명만 매칭
          예) 기준=엔다이브, 규칙키워드="1kg" -> '엔다이브'와 '1kg'가 모두 들어간 상품명만 매칭
             (그래서 다른 상품의 "1kg"는 건드리지 않습니다)

        For mode=map: delta = table[input_qty_key]. If not found and "*" exists, use wildcard.
        If not found and no wildcard -> record missing.
        """
        prod_kw = {p.get("name"): (p.get("keyword") or p.get("name")) for p in cfg.get("products", [])}
        rules_map = cfg.get("rules", {}) or {}

        def _scoped_tokens(base_keyword: str, rule_keyword: str) -> List[str]:
            base_keyword = str(base_keyword or "").strip()
            rule_keyword = str(rule_keyword or "").strip()
            if not rule_keyword:
                return []
            if not base_keyword:
                return [rule_keyword]
            # If user already typed full keyword incl base keyword, treat as absolute substring match
            if base_keyword in rule_keyword:
                return [rule_keyword]
            # Otherwise require BOTH base keyword and rule keyword to exist in product name
            return [base_keyword, rule_keyword]

        actions_raw: List[Dict[str, Any]] = []
        missing: List[Dict[str, Any]] = []

        for base, qty in inputs.items():
            qty = float(qty or 0.0)
            if qty == 0:
                continue

            base_kw = prod_kw.get(base, base)

            rule_list = rules_map.get(base, [])
            if rule_list:
                for r in rule_list:
                    rr = Rule.from_dict(r)
                    if not rr.keyword:
                        continue

                    if rr.mode == "map":
                        k = qty_key(qty)
                        if k in rr.table:
                            delta = rr.table[k]
                        elif "*" in rr.table:
                            delta = rr.table["*"]
                        else:
                            missing.append({"기준상품": base, "키워드": rr.keyword, "입력값": qty, "사유": f"매핑에 '{k}' 없음"})
                            continue
                    else:  # mul
                        delta = qty * rr.value

                    tokens = _scoped_tokens(base_kw, rr.keyword)
                    if not tokens:
                        continue

                    display = " & ".join(tokens) if len(tokens) > 1 else tokens[0]

                    actions_raw.append({
                        "base": base,
                        "match_tokens": tokens,
                        "display": display,
                        "delta": float(delta),
                    })
            else:
                # No rule: base keyword alone is used (legacy behavior)
                tokens = [str(base_kw)]
                actions_raw.append({"base": base, "match_tokens": tokens, "display": tokens[0], "delta": float(qty)})

        # merge by match_tokens key (so different base products with same suffix like "1kg" won't collide)
        merged: Dict[str, float] = {}
        bases: Dict[str, set] = {}
        meta: Dict[str, Dict[str, Any]] = {}

        for a in actions_raw:
            key = "\u0001".join(a["match_tokens"])
            merged[key] = merged.get(key, 0.0) + float(a["delta"])
            bases.setdefault(key, set()).add(a["base"])
            if key not in meta:
                meta[key] = {"match_tokens": a["match_tokens"], "display": a.get("display")}

        out = []
        for k, v in merged.items():
            out.append({
                "match_tokens": meta[k]["match_tokens"],
                "display": meta[k].get("display") or " & ".join(meta[k]["match_tokens"]),
                "delta": v,
                "bases": sorted(list(bases.get(k, []))),
            })

        out.sort(key=lambda x: len(str(x.get("display", ""))), reverse=True)
        return out, pd.DataFrame(missing)


    def embed_config_into_workbook(wb, cfg: Dict[str, Any]) -> None:
        """Save cfg JSON into a hidden sheet so settings can travel with the Excel file."""
        if EXCEL_CONFIG_SHEET in wb.sheetnames:
            ws_cfg = wb[EXCEL_CONFIG_SHEET]
        else:
            ws_cfg = wb.create_sheet(EXCEL_CONFIG_SHEET)
        ws_cfg[EXCEL_CONFIG_CELL].value = json.dumps(cfg, ensure_ascii=False)
        try:
            ws_cfg.sheet_state = "hidden"
        except Exception:
            pass


    def extract_config_from_workbook_bytes(xlsx_bytes: bytes) -> Dict[str, Any]:
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        if EXCEL_CONFIG_SHEET not in wb.sheetnames:
            raise ValueError("엑셀에 저장된 설정 시트가 없습니다.")
        ws_cfg = wb[EXCEL_CONFIG_SHEET]
        raw = ws_cfg[EXCEL_CONFIG_CELL].value
        if not raw:
            raise ValueError("엑셀 설정 셀(A1)이 비어있습니다.")
        cfg = json.loads(str(raw))
        if not isinstance(cfg, dict):
            raise ValueError("엑셀 설정 형식이 올바르지 않습니다.")
        cfg.setdefault("version", 9)
        cfg.setdefault("inventory_column", "재고수량")
        cfg.setdefault("name_column", "상품명")
        cfg.setdefault("products", [])
        cfg.setdefault("rules", {})
        return cfg




    def update_workbook_bytes(
        xlsx_bytes: bytes,
        cfg: Dict[str, Any],
        inputs: Dict[str, float],
    ) -> Tuple[bytes, pd.DataFrame, pd.DataFrame, bytes]:
        """
        Returns:
          - updated workbook bytes
          - df_changes (summary)
          - df_missing (map missing table)
          - changed_rows_only workbook bytes (header + changed rows)
        """
        inv_col = cfg.get("inventory_column", "재고수량")
        name_col = cfg.get("name_column", "상품명")

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active

        header_row, name_idx, inv_idx = find_header_row_and_columns(ws, name_col=name_col, inv_col=inv_col)

        actions, df_missing = build_actions(cfg, inputs)

        def match_action(a: Dict[str, Any], name_str: str) -> bool:
            toks = a.get("match_tokens") or []
            return bool(toks) and all((t in name_str) for t in toks)

        changes = []
        changed_row_indices: List[int] = []

        for r in range(header_row + 1, ws.max_row + 1):
            name_val = ws.cell(r, name_idx).value
            if name_val is None or str(name_val).strip() == "":
                continue

            # skip guideline rows ("필수")
            first_cell = ws.cell(r, 1).value
            if isinstance(first_cell, str) and first_cell.strip() in ("필수",):
                continue
            if isinstance(name_val, str) and name_val.strip() in ("필수",):
                continue

            name_str = str(name_val)
            matched = [a for a in actions if match_action(a, name_str)]
            if not matched:
                continue

            delta = sum(float(m["delta"]) for m in matched)
            if abs(delta) < 1e-12:
                continue

            old = to_number(ws.cell(r, inv_idx).value)
            new = old + delta
            ws.cell(r, inv_idx).value = new

            # record changed row index (to keep same format, we will delete other rows later)
            changed_row_indices.append(r)

            changes.append({
                "행번호": r,
                "상품명": name_str,
                "기존재고": old,
                "증감": delta,
                "최종재고": new,
                "매칭키워드": ", ".join([str(m.get("display") or " & ".join(m.get("match_tokens", []))) for m in matched]),
                "원천상품": ", ".join(sorted({b for m in matched for b in m.get("bases", [])})),
            })

        # Embed current config into output Excel (portable persistence)
        embed_config_into_workbook(wb, cfg)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        updated_bytes = out.getvalue()
        # Build "changed rows only" workbook while keeping the same format/spec as the uploaded Excel:
        # 1) reload the UPDATED workbook bytes (so styles/column widths/etc. are preserved)
        # 2) delete all non-changed data rows (below header_row) from bottom to top
        wb2 = load_workbook(io.BytesIO(updated_bytes))
        ws2 = wb2.active

        keep = set(changed_row_indices)
        # delete from bottom to avoid shifting
        for rr in range(ws2.max_row, header_row, -1):
            if rr not in keep:
                ws2.delete_rows(rr, 1)

        out2 = io.BytesIO()
        wb2.save(out2)
        out2.seek(0)
        changed_rows_bytes = out2.getvalue()

        return updated_bytes, pd.DataFrame(changes), df_missing, changed_rows_bytes


    # ============================
    # Auto mode helpers
    # ============================
    RELAY_BASE_URL = (os.environ.get("RELAY_BASE_URL") or "").strip().rstrip("/")
    RELAY_SHARED_TOKEN = (os.environ.get("RELAY_SHARED_TOKEN") or "").strip()
    TELEGRAM_BOT_TOKEN = (os.environ.get("TELEGRAM_BOT_TOKEN") or "").strip()
    TELEGRAM_CHAT_ID = (os.environ.get("TELEGRAM_CHAT_ID") or "").strip()
    TELEGRAM_AUTO_POLL_SECONDS = max(2, int((os.environ.get("TELEGRAM_AUTO_POLL_SECONDS") or "5").strip() or "5"))
    TELEGRAM_AUTO_TIMEOUT_SECONDS = max(60, int((os.environ.get("TELEGRAM_AUTO_TIMEOUT_SECONDS") or "1200").strip() or "1200"))

    def _safe_json(resp):
        try:
            return resp.json()
        except Exception:
            return {"text": getattr(resp, "text", "")}

    def _get_required_auto_env_missing() -> List[str]:
        missing = []
        for k, v in [
            ("RELAY_BASE_URL", RELAY_BASE_URL),
            ("RELAY_SHARED_TOKEN", RELAY_SHARED_TOKEN),
            ("TELEGRAM_BOT_TOKEN", TELEGRAM_BOT_TOKEN),
            ("TELEGRAM_CHAT_ID", TELEGRAM_CHAT_ID),
        ]:
            if not str(v or "").strip():
                missing.append(k)
        return missing

    def _get_search_body(page: int, size: int) -> Dict[str, Any]:
        raw = (os.environ.get("NAVER_PRODUCT_SEARCH_BODY") or "").strip()
        body: Dict[str, Any] = {}
        if raw:
            try:
                loaded = json.loads(raw)
                if isinstance(loaded, dict):
                    body.update(loaded)
            except Exception:
                pass
        body.setdefault("page", page)
        body.setdefault("size", size)
        return body

    def _iter_dicts(node: Any):
        if isinstance(node, dict):
            yield node
            for v in node.values():
                yield from _iter_dicts(v)
        elif isinstance(node, list):
            for item in node:
                yield from _iter_dicts(item)

    def _relay_request(path: str, *, json_body: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        if not RELAY_BASE_URL:
            raise RuntimeError("RELAY_BASE_URL 환경변수가 비어 있습니다.")
        if not RELAY_SHARED_TOKEN:
            raise RuntimeError("RELAY_SHARED_TOKEN 환경변수가 비어 있습니다.")
        url = f"{RELAY_BASE_URL}{path}"
        resp = requests.post(
            url,
            headers={
                "Authorization": f"Bearer {RELAY_SHARED_TOKEN}",
                "Content-Type": "application/json",
            },
            json=json_body or {},
            timeout=60,
        )
        data = _safe_json(resp)
        if resp.status_code >= 400:
            raise RuntimeError(f"중계서버 호출 실패: POST {path} / {resp.status_code} / {data}")
        if not isinstance(data, dict):
            raise RuntimeError(f"중계서버 응답 형식이 올바르지 않습니다: {data}")
        if data.get("ok") is False:
            raise RuntimeError(f"중계서버 처리 실패: {data.get('status_code')} / {data.get('text') or data.get('data')}")
        return data

    def _extract_products_from_search_payload(payload: Any) -> pd.DataFrame:
        rows: List[Dict[str, Any]] = []
        seen = set()
        for d in _iter_dicts(payload):
            cps = d.get("channelProducts")
            if isinstance(cps, list) and cps:
                origin_no = d.get("originProductNo")
                for cp in cps:
                    if not isinstance(cp, dict):
                        continue
                    row = {
                        "originProductNo": cp.get("originProductNo") or origin_no,
                        "channelProductNo": cp.get("channelProductNo") or cp.get("smartstoreChannelProductNo") or cp.get("id"),
                        "name": cp.get("name") or cp.get("channelProductName") or d.get("name") or d.get("originProductName"),
                        "stockQuantity": cp.get("stockQuantity"),
                    }
                    try:
                        row["originProductNo"] = int(row["originProductNo"]) if row["originProductNo"] is not None else None
                    except Exception:
                        row["originProductNo"] = None
                    try:
                        row["stockQuantity"] = float(row["stockQuantity"]) if row["stockQuantity"] is not None else 0.0
                    except Exception:
                        row["stockQuantity"] = 0.0
                    key = (row.get("originProductNo"), row.get("channelProductNo"), row.get("name"))
                    if row.get("originProductNo") is not None and row.get("name") and key not in seen:
                        seen.add(key)
                        rows.append(row)
            elif d.get("originProductNo") is not None and (d.get("stockQuantity") is not None) and (d.get("name") or d.get("originProductName")):
                row = {
                    "originProductNo": d.get("originProductNo"),
                    "channelProductNo": d.get("channelProductNo") or d.get("smartstoreChannelProductNo") or d.get("id"),
                    "name": d.get("name") or d.get("originProductName"),
                    "stockQuantity": d.get("stockQuantity"),
                }
                try:
                    row["originProductNo"] = int(row["originProductNo"]) if row["originProductNo"] is not None else None
                except Exception:
                    row["originProductNo"] = None
                try:
                    row["stockQuantity"] = float(row["stockQuantity"]) if row["stockQuantity"] is not None else 0.0
                except Exception:
                    row["stockQuantity"] = 0.0
                key = (row.get("originProductNo"), row.get("channelProductNo"), row.get("name"))
                if row.get("originProductNo") is not None and row.get("name") and key not in seen:
                    seen.add(key)
                    rows.append(row)
        if not rows:
            return pd.DataFrame(columns=["originProductNo", "channelProductNo", "name", "stockQuantity"])
        df = pd.DataFrame(rows)
        df["name"] = df["name"].astype(str).str.strip()
        df = df[df["name"] != ""].copy()
        return df[["originProductNo", "channelProductNo", "name", "stockQuantity"]].drop_duplicates(subset=["originProductNo", "channelProductNo", "name"])

    def fetch_naver_products_df() -> pd.DataFrame:
        relay_resp = _relay_request("/naver/products/search", json_body={"body": _get_search_body(page=1, size=500)})
        payload = relay_resp.get("data") or {}
        df = _extract_products_from_search_payload(payload)
        if df.empty:
            raise RuntimeError("중계서버를 통해 받은 네이버 상품 조회 결과가 없습니다. NAVER_PRODUCT_SEARCH_BODY 또는 중계서버 응답을 확인해 주세요.")
        return df.drop_duplicates(subset=["originProductNo", "channelProductNo", "name"]).reset_index(drop=True)

    def compute_stock_display_map_from_df(df_products: pd.DataFrame, cfg: Dict[str, Any]) -> Dict[str, str]:
        """
        중계서버에서 받은 상품 DataFrame에도 수동 모드와 같은 방식으로
        상품목록관리/규칙관리 기준 재고 표시를 적용합니다.
        """
        out: Dict[str, str] = {}
        if df_products is None or df_products.empty:
            return out

        recog_logic = _get_recognition_logic(cfg)
        products = cfg.get("products", []) or []
        base_kw: Dict[str, str] = {}
        for p in products:
            bn = str(p.get("name", "")).strip()
            kw = str(p.get("keyword") or p.get("name") or "").strip()
            if bn:
                base_kw[bn] = kw or bn

        bases_sorted = sorted(
            [(bn, base_kw.get(bn, bn)) for bn in base_kw.keys()],
            key=lambda x: len(str(x[1] or "")),
            reverse=True,
        )
        if not bases_sorted:
            return out

        rules_map = cfg.get("rules", {}) or {}

        def _scoped_tokens(base_keyword: str, rule_keyword: str) -> List[str]:
            base_keyword = str(base_keyword or "").strip()
            rule_keyword = str(rule_keyword or "").strip()
            if not rule_keyword:
                return []
            if not base_keyword:
                return [rule_keyword]
            if base_keyword in rule_keyword:
                return [rule_keyword]
            return [base_keyword, rule_keyword]

        def _match_tokens(tokens: List[str], name_str: str) -> bool:
            return bool(tokens) and all((t in name_str) for t in tokens)

        matchers_by_base: Dict[str, List[Dict[str, Any]]] = {}
        unit_pref: Dict[str, str] = {}
        unit_seen: Dict[str, str] = {}

        for bn, bkw in bases_sorted:
            rs = rules_map.get(bn, []) or []
            ms: List[Dict[str, Any]] = []
            unit_counts: Dict[str, int] = {}

            for i, r in enumerate(rs):
                rr = Rule.from_dict(r)
                if not rr.keyword:
                    continue
                tokens = _scoped_tokens(bkw, rr.keyword)
                if not tokens:
                    continue
                factor, unit = _parse_factor_and_unit(rr.keyword, recog_logic)
                ms.append({
                    "tokens": tokens,
                    "keyword": rr.keyword,
                    "factor": factor,
                    "unit": unit,
                    "order": i,
                })
                if unit:
                    unit_counts[unit] = unit_counts.get(unit, 0) + 1

            ms.sort(
                key=lambda m: (sum(len(t) for t in m["tokens"]), len(m["tokens"]), len(m["keyword"])),
                reverse=True,
            )
            matchers_by_base[bn] = ms

            if unit_counts.get("kg", 0) > 0:
                unit_pref[bn] = "kg"
            elif unit_counts:
                unit_pref[bn] = sorted(unit_counts.items(), key=lambda kv: (kv[1], len(kv[0])), reverse=True)[0][0]
            else:
                unit_pref[bn] = ""

        totals: Dict[str, Decimal] = {}

        for _, row in df_products.iterrows():
            name_str = str(row.get("name") or "").strip()
            if not name_str:
                continue

            base_name = None
            for bn, kw in bases_sorted:
                if kw and kw in name_str:
                    base_name = bn
                    break
            if base_name is None:
                continue

            inv_qty = _cell_to_decimal(row.get("stockQuantity"))
            if inv_qty == 0:
                continue

            chosen = None
            for m in matchers_by_base.get(base_name, []):
                if _match_tokens(m["tokens"], name_str):
                    chosen = m
                    break

            has_rules = bool(matchers_by_base.get(base_name))
            if chosen and chosen.get("unit"):
                factor = chosen["factor"]
                unit = chosen["unit"]
            else:
                if has_rules and chosen is None:
                    continue
                factor, unit = _parse_factor_and_unit(name_str, recog_logic)

            pref = unit_pref.get(base_name, "") or ""
            if pref:
                unit_seen[base_name] = pref
            elif unit:
                unit_seen[base_name] = unit_seen.get(base_name) or unit

            totals[base_name] = totals.get(base_name, Decimal("0")) + (inv_qty * factor)

        for bn, _kw in bases_sorted:
            total = totals.get(bn, Decimal("0"))
            if total == 0:
                out[bn] = "0"
                continue
            unit = unit_pref.get(bn, "") or unit_seen.get(bn, "") or ""
            if unit == "kg":
                out[bn] = f"{_fmt_decimal(total, max_decimals=3)}kg"
            else:
                s = _fmt_decimal(total, max_decimals=3)
                out[bn] = f"{s}{unit}" if unit else s
        return out

    def apply_auto_stock_to_df(df_work: pd.DataFrame, df_products: pd.DataFrame, cfg: Dict[str, Any]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        df2 = df_work.copy()
        current_rows = []
        for _, row in df_products.iterrows():
            current_rows.append({
                "originProductNo": row.get("originProductNo"),
                "channelProductNo": row.get("channelProductNo"),
                "실제상품명": row.get("name"),
                "현재재고": float(row.get("stockQuantity") or 0),
            })
        current_df = pd.DataFrame(current_rows)
        changes: List[Dict[str, Any]] = []
        missing: List[Dict[str, Any]] = []

        for i, work_row in df2.iterrows():
            product = str(work_row.get("상품") or "").strip()
            try:
                input_qty = float(work_row.get("입력수량") or 0)
            except Exception:
                input_qty = 0.0
            if input_qty <= 0:
                continue
            cfg_row = None
            for p in cfg.get("products", []):
                if str(p.get("name") or "").strip() == product:
                    cfg_row = p
                    break
            pattern = str((cfg_row or {}).get("pattern") or (cfg_row or {}).get("keyword") or product).strip()
            matched = current_df[current_df["실제상품명"].astype(str).str.contains(re.escape(pattern), na=False)] if pattern else current_df.iloc[0:0]
            if matched.empty:
                missing.append({"상품": product, "패턴": pattern, "사유": "매칭 실패"})
                continue
            target = matched.iloc[0]
            before_qty = float(target.get("현재재고") or 0)
            after_qty = before_qty + input_qty
            df2.at[i, "재고수량"] = fmt_qty_no_zero(after_qty)
            changes.append({
                "상품": product,
                "입력수량": input_qty,
                "현재재고": before_qty,
                "최종재고": after_qty,
                "originProductNo": int(target.get("originProductNo")),
                "channelProductNo": target.get("channelProductNo"),
                "실제상품명": str(target.get("실제상품명") or ""),
                "패턴": pattern,
            })
        return df2, pd.DataFrame(changes), pd.DataFrame(missing)

    def build_multi_update_payload(df_changes: pd.DataFrame) -> Dict[str, Any]:
        items: List[Dict[str, Any]] = []
        if df_changes is None or df_changes.empty:
            return {"items": []}
        for _, row in df_changes.iterrows():
            try:
                qty = int(round(float(row.get("최종재고", 0))))
                origin_no = int(row.get("originProductNo"))
            except Exception:
                continue
            if qty <= 0:
                continue
            items.append({"originProductNo": origin_no, "stockQuantity": qty})
        return {"items": items}

    def push_stock_updates(df_changes: pd.DataFrame) -> Dict[str, Any]:
        payload = build_multi_update_payload(df_changes)
        if not payload.get("items"):
            return {"sent": 0, "response": {"message": "변경 대상 없음"}, "payload": payload}
        data = _relay_request("/naver/stock/update", json_body=payload)
        try:
            sent_count = int(data.get("sent_count") or len(payload.get("items") or []))
        except Exception:
            sent_count = len(payload.get("items") or [])
        return {"sent": sent_count, "response": data.get("data") or data, "payload": payload, "relay": data}

    def _telegram_request(method: str, *, params: Optional[Dict[str, Any]] = None, json_body: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        if not TELEGRAM_BOT_TOKEN:
            raise RuntimeError("TELEGRAM_BOT_TOKEN 환경변수가 비어 있습니다.")
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/{method}"
        resp = requests.post(url, params=params, json=json_body, timeout=30)
        data = _safe_json(resp)
        if resp.status_code >= 400 or not data.get("ok", False):
            raise RuntimeError(f"텔레그램 API 실패: {method} / {resp.status_code} / {data}")
        return data

    def send_telegram_message(text: str) -> Dict[str, Any]:
        if not TELEGRAM_CHAT_ID:
            raise RuntimeError("TELEGRAM_CHAT_ID 환경변수가 비어 있습니다.")
        return _telegram_request("sendMessage", json_body={"chat_id": TELEGRAM_CHAT_ID, "text": text})

    def get_latest_valid_reply(after_unix_ts: int) -> Optional[Dict[str, Any]]:
        offset = int(st.session_state.get("bulk_tg_update_offset") or 0)
        data = _telegram_request("getUpdates", params={"timeout": 0, "offset": offset + 1})
        result = data.get("result") or []
        latest = None
        max_update_id = offset
        for item in result:
            update_id = int(item.get("update_id") or 0)
            max_update_id = max(max_update_id, update_id)
            msg = item.get("message") or item.get("edited_message") or {}
            if not msg:
                continue
            if str(msg.get("chat", {}).get("id")) != str(TELEGRAM_CHAT_ID):
                continue
            date_ts = int(msg.get("date") or 0)
            text = str(msg.get("text") or "").strip()
            if date_ts < int(after_unix_ts):
                continue
            if not text:
                continue
            latest = {"update_id": update_id, "date": date_ts, "text": text}
            break
        st.session_state["bulk_tg_update_offset"] = max_update_id
        return latest

    def build_current_stock_message(current_names: List[str], stock_map: Dict[str, str]) -> str:
        lines = ["[현재 재고 수량]"]
        for i, name in enumerate(current_names, start=1):
            lines.append(f"{i}. {name} {stock_map.get(name, '0')}")
        return "\n".join(lines)

    def build_input_template_message(current_names: List[str]) -> str:
        lines = []
        for i, _name in enumerate(current_names, start=1):
            lines.append(f"{i} 0")
        return "\n".join(lines)

    def build_invalid_reply_message(current_names: List[str], error_text: str) -> str:
        lines = ["[입력 형식 오류]", str(error_text).strip(), "", "아래 형식으로 다시 보내주세요."]
        sample_count = min(5, len(current_names))
        for i in range(1, sample_count + 1):
            lines.append(f"{i} 0")
        lines.append("")
        lines.append("전체를 다시 보내거나, '취소'를 보내면 자동 실행을 중지합니다.")
        return "\n".join(lines)

    def parse_number_qty_reply(text: str, max_index: int) -> Tuple[bool, Dict[int, float], str]:
        parsed: Dict[int, float] = {}
        lines = [ln.strip() for ln in str(text).splitlines() if ln.strip()]
        if not lines:
            return False, {}, "메시지가 비어 있습니다."
        for i, line in enumerate(lines, start=1):
            m = re.fullmatch(r"(\d+)\s+([0-9]+(?:\.[0-9]+)?)", line)
            if not m:
                return False, {}, f"{i}번째 줄 형식이 올바르지 않습니다. 예: 1 5"
            idx = int(m.group(1))
            qty = float(m.group(2))
            if idx < 1 or idx > max_index:
                return False, {}, f"{i}번째 줄 상품 번호가 범위를 벗어났습니다: {idx}"
            if qty < 0:
                return False, {}, f"{i}번째 줄 수량은 음수일 수 없습니다."
            parsed[idx] = parsed.get(idx, 0.0) + qty
        return True, parsed, ""

    def build_auto_inputs(current_names: List[str], parsed_reply: Dict[int, float]) -> Dict[str, float]:
        out: Dict[str, float] = {name: 0.0 for name in current_names}
        for idx, qty in parsed_reply.items():
            if 1 <= idx <= len(current_names):
                out[current_names[idx - 1]] = float(qty)
        return out

    def build_auto_memo_text(current_names: List[str], inputs: Dict[str, float]) -> str:
        return "\n".join([f"{name} - {fmt_qty_for_memo(float(inputs.get(name, 0.0) or 0.0))}" for name in current_names])

    def build_confirmation_message(current_names: List[str], inputs: Dict[str, float]) -> str:
        lines = ["[반영 예정 확인]"]
        for i, name in enumerate(current_names, start=1):
            lines.append(f"{i}. {name} {fmt_qty_for_memo(float(inputs.get(name, 0.0) or 0.0))}")
        lines.append("")
        lines.append("맞으면 '확정', 중지하려면 '취소'를 보내주세요.")
        lines.append("수정하려면 템플릿 전체를 다시 보내도 됩니다.")
        return "\n".join(lines)

    def summarize_auto_result(sent_count: int, df_changes: pd.DataFrame, df_missing: pd.DataFrame, api_result: Dict[str, Any]) -> str:
        changed_count = 0 if df_changes is None else int(len(df_changes))
        missing_count = 0 if df_missing is None else int(len(df_missing))
        lines = ["[자동 재고 반영 완료]", f"전송건수 {sent_count}", f"실제 변경행 {changed_count}"]
        if missing_count > 0:
            lines.append(f"미적용 규칙 {missing_count}")
        resp = api_result.get("response") if isinstance(api_result, dict) else None
        if isinstance(resp, dict) and resp.get("message"):
            lines.append(f"응답: {resp.get('message')}")
        return "\n".join(lines)

    def telegram_send_message(text: str) -> Dict[str, Any]:
        return send_telegram_message(text)

    def telegram_get_updates(offset: int = 0) -> List[Dict[str, Any]]:
        data = _telegram_request("getUpdates", params={"timeout": 0, "offset": int(offset)})
        return data.get("result") or []

    def _get_latest_update_id() -> int:
        offset = int(st.session_state.get("bulk_tg_update_offset") or 0)
        updates = telegram_get_updates(offset=offset + 1)
        max_update_id = offset
        for item in updates:
            try:
                update_id = int(item.get("update_id") or 0)
            except Exception:
                continue
            max_update_id = max(max_update_id, update_id)
        st.session_state["bulk_tg_update_offset"] = max_update_id
        return max_update_id

    def parse_first_reply_message(text: str, max_index: int) -> Tuple[bool, Dict[int, float], str]:
        return parse_number_qty_reply(text, max_index)

    def is_cancel_message(text: str) -> bool:
        return str(text or "").strip() == "취소"

    def is_confirm_message(text: str) -> bool:
        return str(text or "").strip() == "확정"

    def build_auto_stop_message(reason: str) -> str:
        reason_text = str(reason or "").strip()
        if not reason_text:
            reason_text = "중지"
        return f"[자동 실행 중지]\n사유: {reason_text}"

    def apply_actions_to_products_df(df_products: pd.DataFrame, cfg: Dict[str, Any], inputs: Dict[str, float]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        actions, df_missing = build_actions(cfg, inputs)
        df_work = df_products.copy() if isinstance(df_products, pd.DataFrame) else pd.DataFrame()
        if df_work.empty:
            return df_work, pd.DataFrame(), df_missing

        def match_action(a: Dict[str, Any], name_str: str) -> bool:
            toks = a.get("match_tokens") or []
            return bool(toks) and all((t in name_str) for t in toks)

        changes: List[Dict[str, Any]] = []
        df_work["name"] = df_work.get("name", "").astype(str)
        for idx, row in df_work.iterrows():
            name_str = str(row.get("name") or "")
            matched = [a for a in actions if match_action(a, name_str)]
            if not matched:
                continue
            delta = sum(float(m.get("delta") or 0) for m in matched)
            if abs(delta) < 1e-12:
                continue
            before_qty = float(row.get("stockQuantity") or 0)
            after_qty = before_qty + delta
            df_work.at[idx, "stockQuantity"] = after_qty
            changes.append({
                "상품명": name_str,
                "기존재고": before_qty,
                "증감": delta,
                "최종재고": after_qty,
                "originProductNo": int(row.get("originProductNo")),
                "channelProductNo": row.get("channelProductNo"),
                "매칭키워드": ", ".join([str(m.get("display") or " & ".join(m.get("match_tokens", []))) for m in matched]),
                "원천상품": ", ".join(sorted({b for m in matched for b in m.get("bases", [])})),
            })
        return df_work, pd.DataFrame(changes), df_missing

    # ============================
    # UI
    # ============================
    cfg = load_config()

    # 자동복원: 상품목록이 비어있으면 기본 34개를 채워넣습니다.
    if not cfg.get("products"):
        cfg["products"] = DEFAULT_PRODUCTS.copy()
        save_config(cfg)

    st.title("🧰 재고일괄변경")

    # ============================
    # Sidebar navigation
    # ============================
    st.sidebar.title("메뉴")
    # ✅ 기본은 접힌 상태(디폴트), 필요할 때만 펼쳐서 메뉴 이동
    with st.sidebar.expander("📂 펼쳐보기", expanded=False):
        page = st.radio(
            "이동",
            options=["① 재고 입력", "② 상품목록 관리", "③ 규칙 관리", "④ 백업/복원"],
            index=0,  # ✅ 시작 페이지: 재고 입력
            key="bulk_stock_sidebar_page",
        )

    # ============================
    # Pages
    # ============================
    if page.startswith("①"):
        mode = st.radio(
            "입력 방식",
            options=["수동", "자동"],
            horizontal=True,
            key="bulk_stock_input_mode",
        )

        if mode == "수동":
            st.subheader("엑셀 파일 불러오기")
            uploaded = st.file_uploader("스마트스토어 수정양식 엑셀(.xlsx)을 업로드하세요", type=["xlsx"], key="bulk_xlsx_uploader")

            # 업로드 바이트(중복 getvalue() 방지)
            uploaded_bytes = uploaded.getvalue() if uploaded is not None else None

            # ✅ '메모장으로 저장' 등으로 rerun 되어도, 마지막 적용 결과(다운로드/변경표)가 사라지지 않게 유지
            if "last_apply_result" not in st.session_state:
                st.session_state.last_apply_result = None

            def _fingerprint_upload(name: str, b: bytes) -> str:
                md5 = hashlib.md5(b).hexdigest()
                return f"{name}:{len(b)}:{md5}"

            current_fp = (
                _fingerprint_upload(getattr(uploaded, "name", "") or "", uploaded_bytes)
                if uploaded_bytes is not None
                else None
            )

            # 다른 엑셀을 새로 업로드하면, 이전 결과는 자동으로 숨김 처리
            _last = st.session_state.get("last_apply_result")
            if _last is not None:
                if (current_fp is None) or (_last.get("fingerprint") != current_fp):
                    st.session_state.last_apply_result = None


            st.subheader("입력할 수량")
            prod_list = cfg.get("products", [])
            if not prod_list:
                st.info("상품목록이 비어있습니다. 사이드바의 '상품목록 관리'에서 상품을 먼저 추가하세요.")
            else:
                # 업로드된 엑셀로부터 '재고수량(표시용)'을 계산 (입력/저장에는 절대 사용하지 않음)
                stock_map: Dict[str, str] = {}
                if uploaded is not None:
                    try:
                        stock_map = compute_stock_display_map(uploaded_bytes, cfg)
                    except Exception:
                        stock_map = {}

                current_names = [str(p.get("name", "")).strip() for p in prod_list if str(p.get("name", "")).strip()]
                if not current_names:
                    st.info("상품목록이 비어있습니다. 사이드바의 '상품목록 관리'에서 상품을 먼저 추가하세요.")
                else:
                                # 입력값은 '✅ 엑셀에 적용하기'를 눌렀을 때만 확정(저장)됩니다.
                    def _align_qty_df(_df: pd.DataFrame, _names: List[str]) -> pd.DataFrame:
                        try:
                            if _df is not None and not _df.empty and "상품" in _df.columns and "입력수량" in _df.columns:
                                _m = _df.set_index("상품")["입력수량"].to_dict()
                            else:
                                _m = {}
                        except Exception:
                            _m = {}

                        return pd.DataFrame(
                            {
                                "상품": _names,
                                "입력수량": ["" if (_m.get(n) is None) else _m.get(n, "") for n in _names],
                            }
                        )

                    # 마지막으로 '적용(확정)'된 값(=엑셀 적용에 사용되는 값)
                    if "qty_committed_df" not in st.session_state:
                        st.session_state.qty_committed_df = pd.DataFrame(
                            {"상품": current_names, "입력수량": [""] * len(current_names)}
                        )
                    st.session_state.qty_committed_df = _align_qty_df(st.session_state.qty_committed_df, current_names)
                    # --- 참고수량: 저장된 값(cfg["ref_qty"])만 표시합니다. (💾 참고수량 저장을 눌러야만 영구 저장됨) ---
                    ref_saved = (cfg.get("ref_qty") or {})

                    def _ref_saved_value(_name: str) -> str:
                        v = ref_saved.get(_name)
                        return "" if v is None else str(v)

                    # 표 표시용(재고수량/참고수량은 표시/입력만)
                    df_view = st.session_state.qty_committed_df.copy()
                    df_view["재고수량"] = [stock_map.get(n, "") for n in df_view["상품"]]
                    df_view["참고수량"] = [_ref_saved_value(n) for n in df_view["상품"]]
                    df_view = df_view[["상품", "입력수량", "재고수량", "참고수량"]]

                    st.caption("※ 표에 값을 입력해도 즉시 저장/적용되지 않습니다. **✅ 엑셀에 적용하기**를 눌러야 엑셀에 반영됩니다.")
                    st.caption("※ '입력수량'을 비워두면 0으로 처리됩니다. (예: 엔다이브 - 0)")
                    st.caption("※ '참고수량'은 **💾 참고수량 저장**을 눌러야 저장되며, 저장 후에는 수정 전까지 유지됩니다.")

                    # 상품목록이 변경되면 편집 상태를 초기화하기 위해 key를 변경합니다.
                    _sig = hashlib.md5(("|".join(current_names)).encode("utf-8")).hexdigest()[:10]
                    _editor_key = f"qty_editor_{_sig}"


                    # data_editor가 rerun 때 값이 사라지는 것을 방지:
                    # - 같은 key의 widget state가 있으면 그 값을 우선 사용
                    # - 재고수량(표시용)만 매번 새로 갱신
                    _df_for_editor = df_view
                    if _editor_key in st.session_state and isinstance(st.session_state.get(_editor_key), pd.DataFrame):
                        _prev = st.session_state.get(_editor_key).copy()
                        try:
                            if "상품" in _prev.columns:
                                _prev["상품"] = _prev["상품"].astype(str)
                                _prev = _prev.set_index("상품").reindex(current_names).reset_index()
                        except Exception:
                            _prev = df_view.copy()

                        for _c in ["입력수량", "재고수량", "참고수량"]:
                            if _c not in _prev.columns:
                                _prev[_c] = ""

                        _prev["재고수량"] = [stock_map.get(n, "") for n in _prev["상품"]]
                        _df_for_editor = _prev[["상품", "입력수량", "재고수량", "참고수량"]]

                    df_edit = st.data_editor(
                        _df_for_editor,
                        key=_editor_key,
                        use_container_width=True,
                        num_rows="fixed",
                        disabled=["상품", "재고수량"],
                        column_config={
                            "입력수량": st.column_config.TextColumn("입력수량", help="숫자 입력(음수/소수 가능). 예: 3, -2, 1.5"),
                            "재고수량": st.column_config.TextColumn("재고수량", help="업로드한 엑셀 기준, 모든 옵션 재고 합(표시용)"),
                            "참고수량": st.column_config.TextColumn("참고수량", help="메모/참고용 수량(저장 시 유지). 예: 10"),
                        },
                    )


                    # 참고수량은 표에서 편집할 수 있지만, **💾 참고수량 저장**을 눌러야만 설정에 저장됩니다.

                    # ✅ 현재 표(편집중 값) 기준으로 입력값/메모 생성 (빈칸은 0)
                    df_inputs = df_edit.drop(columns=["재고수량", "참고수량"], errors="ignore").copy()

                    inputs: Dict[str, float] = {}
                    memo_lines: List[str] = []
                    for _, r in df_inputs.iterrows():
                        name = str(r.get("상품", "")).strip()
                        if not name:
                            continue
                        qty = parse_input_number(r.get("입력수량", ""))
                        inputs[name] = qty
                        memo_lines.append(f"{name} - {fmt_qty_for_memo(qty)}")
                    memo_text = "\n".join(memo_lines)

                    col_a, col_b, col_c = st.columns(3, gap="small")
                    with col_a:
                        apply_clicked = st.button(
                            "✅ 엑셀에 적용하기",
                            disabled=(uploaded is None),
                            use_container_width=True,
                        )
                    with col_b:
                        memo_filename = "주문내역.txt"
                        st.download_button(
                            "📝 메모장으로 저장",
                            data=memo_text.encode("utf-8"),
                            file_name=memo_filename,
                            mime="text/plain",
                            disabled=(len(memo_lines) == 0),
                            help="현재 표(편집중 값) 기준으로 텍스트(.txt)를 다운로드합니다. (빈칸=0)",
                            use_container_width=True,
                        )
                    with col_c:
                        ref_save_clicked = st.button(
                            "💾 참고수량 저장",
                            use_container_width=True,
                            key=f"bulk_save_ref_qty_{_editor_key}",
                        )

                    # 💾 참고수량 저장 처리(저장 후 수정 전까지 유지)
                    if ref_save_clicked:
                        ref_map_new: Dict[str, Any] = {}
                        for _, rr in df_edit.iterrows():
                            pname = str(rr.get("상품", "")).strip()
                            if not pname:
                                continue

                            raw = rr.get("참고수량", "")
                            if raw is None:
                                sval = ""
                            elif isinstance(raw, (int, float)):
                                sval = fmt_qty_for_memo(float(raw))
                            else:
                                sval = str(raw).strip()
                                if sval != "":
                                    s_clean = sval.replace(",", "")
                                    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?", s_clean):
                                        sval = fmt_qty_for_memo(parse_input_number(sval))

                            if sval == "":
                                continue
                            ref_map_new[pname] = sval

                        cfg["ref_qty"] = ref_map_new
                        save_config(cfg)
                        st.success("참고수량 저장 완료!")
                        st.rerun()

                    if apply_clicked:
                        # ✅ 적용 버튼을 눌렀을 때만 '확정(저장)' + 엑셀 반영
                        st.session_state.qty_committed_df = df_inputs.copy()

                        if uploaded_bytes is None:
                            st.warning("엑셀 파일을 업로드해 주세요.")
                            st.session_state.last_apply_result = None
                        else:
                            try:
                                updated_bytes, df_changes, df_missing, changed_rows_bytes = update_workbook_bytes(
                                    uploaded_bytes, cfg, inputs
                                )
                            except Exception as e:
                                st.error(f"적용 중 오류: {e}")
                                st.session_state.last_apply_result = None
                            else:
                                out_name_changed = "재고수량일괄변경.xlsx"

                                # ✅ 결과를 session_state에 저장 (메모장 다운로드 등 rerun에도 유지)
                                st.session_state.last_apply_result = {
                                    "fingerprint": current_fp,
                                    "out_name": out_name_changed,
                                    "changed_rows_bytes": changed_rows_bytes,
                                    "df_changes": df_changes,
                                    "df_missing": df_missing,
                                    "applied_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                }

                    # ✅ 마지막 적용 결과 표시(버튼/표가 rerun으로 사라지지 않음)
                    _last = st.session_state.get("last_apply_result")
                    if _last is not None:
                        st.success(f"완료! 아래에서 다운로드하세요. (마지막 적용: {_last.get('applied_at', '')})")

                        st.download_button(
                            "⬇️ 다운로드",
                            data=_last["changed_rows_bytes"],
                            file_name=_last["out_name"],
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            help="재고수량이 변경된 행(헤더 포함)만 남긴 파일입니다.",
                            use_container_width=True,
                            key=f"bulk_excel_dl_{_last.get('fingerprint', '')}",
                        )

                        df_changes = _last.get("df_changes")
                        df_missing = _last.get("df_missing")

                        if isinstance(df_changes, pd.DataFrame):
                            if df_changes.empty:
                                st.info("변경된 행이 없습니다. (키워드 매칭이 안 됐거나 입력이 0이거나, map 모드에서 입력값이 매핑에 없을 수 있어요)")
                            else:
                                st.dataframe(df_changes.drop(columns=["행번호"], errors="ignore"), use_container_width=True)

                        if isinstance(df_missing, pd.DataFrame) and (not df_missing.empty):
                            st.warning("⚠️ map(매핑) 규칙에서 '입력값 → 적용값'이 정의되지 않아 적용되지 않은 항목이 있습니다.")
                            st.dataframe(df_missing, use_container_width=True)

        else:
            st.subheader("자동 실행")
            st.caption("시작 버튼을 누르면 텔레그램으로 0 템플릿을 먼저 보내고, 그 다음 현재 재고 수량을 보낸 뒤 확정/취소를 기다립니다.")

            auto_state = st.session_state.get("bulk_auto_state")
            if not isinstance(auto_state, dict):
                auto_state = {
                    "status": "idle",
                    "polling": False,
                    "error": "",
                    "message_text": "",
                    "reply_text": "",
                    "memo_text": "",
                    "summary_text": "",
                    "confirm_text": "",
                    "base_update_id": 0,
                    "sent_at": 0,
                    "names": [],
                    "products_records": [],
                    "processed_update_id": None,
                    "last_check_at": "",
                    "run_id": "",
                    "deadline_ts": 0,
                    "pending_inputs": {},
                }
                st.session_state.bulk_auto_state = auto_state

            col1, col2 = st.columns(2, gap="small")
            with col1:
                start_auto = st.button("▶ 자동 실행 시작", use_container_width=True, key="bulk_auto_start")
            with col2:
                stop_auto = st.button("⏹ 자동 실행 중지", use_container_width=True, key="bulk_auto_stop")

            if stop_auto:
                auto_state["polling"] = False
                auto_state["status"] = "stopped"
                auto_state["summary_text"] = build_auto_stop_message("수동 중지")
                st.session_state.bulk_auto_state = auto_state
                try:
                    telegram_send_message(auto_state["summary_text"])
                except Exception:
                    pass
                st.info("자동 실행을 중지했습니다.")

            if start_auto:
                missing_env = _get_required_auto_env_missing()
                if missing_env:
                    auto_state.update({"status": "error", "polling": False, "error": "필수 환경변수가 없습니다: " + ", ".join(missing_env)})
                    st.session_state.bulk_auto_state = auto_state
                else:
                    try:
                        df_products = fetch_naver_products_df()
                        stock_map = compute_stock_display_map_from_df(df_products, cfg)
                        current_names = [str(p.get("name", "")).strip() for p in cfg.get("products", []) if str(p.get("name", "")).strip()]
                        base_update_id = _get_latest_update_id()
                        message_text = build_current_stock_message(current_names, stock_map)
                        template_text = build_input_template_message(current_names)
                        telegram_send_message(template_text)
                        telegram_send_message(message_text)
                    except Exception as e:
                        auto_state.update({"status": "error", "polling": False, "error": str(e)})
                        st.session_state.bulk_auto_state = auto_state
                    else:
                        now_ts = int(time.time())
                        auto_state = {
                            "status": "waiting_reply",
                            "polling": True,
                            "error": "",
                            "message_text": message_text,
                            "template_text": template_text,
                            "reply_text": "",
                            "memo_text": "",
                            "summary_text": "",
                            "confirm_text": "",
                            "base_update_id": int(base_update_id),
                            "sent_at": now_ts,
                            "names": current_names,
                            "products_records": df_products.to_dict(orient="records"),
                            "processed_update_id": None,
                            "last_check_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "run_id": now_prefix_kst(),
                            "deadline_ts": now_ts + TELEGRAM_AUTO_TIMEOUT_SECONDS,
                            "pending_inputs": {},
                        }
                        st.session_state.bulk_auto_state = auto_state
                        st.success("텔레그램으로 현재 재고/입력 템플릿을 전송했습니다. 답장을 기다립니다.")
                        st.rerun()

            auto_state = st.session_state.get("bulk_auto_state") or {}
            status = auto_state.get("status", "idle")
            poll_active = bool(auto_state.get("polling"))

            if status == "error" and auto_state.get("error"):
                st.error(auto_state.get("error"))
            elif status == "done":
                st.success("자동 반영이 완료되었습니다.")
            elif status == "stopped":
                st.info("자동 실행이 중지되었습니다.")
            elif status == "waiting_confirm":
                st.info("입력값을 받았습니다. 텔레그램에서 '확정' 또는 '취소'를 기다리는 중입니다.")
            elif poll_active:
                st.info("텔레그램 답장을 기다리는 중입니다.")
            else:
                st.caption("자동 실행 시작을 누르면 텔레그램 답장을 자동 확인합니다.")

            if auto_state.get("message_text"):
                st.text(auto_state.get("message_text"))
            if auto_state.get("reply_text"):
                st.caption("최근 입력 답장")
                st.text(auto_state.get("reply_text"))
            if auto_state.get("confirm_text"):
                st.caption("확정 대기 메시지")
                st.text(auto_state.get("confirm_text"))
            if auto_state.get("memo_text"):
                st.caption("메모장 텍스트")
                st.text(auto_state.get("memo_text"))
            if auto_state.get("summary_text"):
                st.caption("최종 결과")
                st.text(auto_state.get("summary_text"))
            if auto_state.get("last_check_at"):
                st.caption(f"마지막 확인: {auto_state.get('last_check_at')}")
            if auto_state.get("deadline_ts") and poll_active:
                remain = max(0, int(auto_state.get("deadline_ts") or 0) - int(time.time()))
                st.caption(f"남은 대기시간: {remain // 60}분 {remain % 60}초")

            if poll_active and status in ("waiting_reply", "waiting_confirm"):
                try:
                    now_ts = int(time.time())
                    if int(auto_state.get("deadline_ts") or 0) > 0 and now_ts > int(auto_state.get("deadline_ts") or 0):
                        stop_text = build_auto_stop_message("20분 동안 답장이 없어 자동 실행을 중지했습니다.")
                        auto_state["status"] = "stopped"
                        auto_state["polling"] = False
                        auto_state["summary_text"] = stop_text
                        st.session_state.bulk_auto_state = auto_state
                        telegram_send_message(stop_text)
                        st.rerun()

                    updates = telegram_get_updates(offset=int(auto_state.get("base_update_id", 0)) + 1)
                    auto_state["last_check_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    for upd in updates:
                        try:
                            update_id = int(upd.get("update_id") or 0)
                        except Exception:
                            continue
                        msg = upd.get("message") or upd.get("edited_message") or {}
                        chat = msg.get("chat") or {}
                        if str(chat.get("id") or "") != str(TELEGRAM_CHAT_ID):
                            continue
                        if int(msg.get("date") or 0) < int(auto_state.get("sent_at") or 0):
                            continue
                        text_msg = str(msg.get("text") or "").strip()
                        if not text_msg:
                            continue
                        auto_state["processed_update_id"] = update_id
                        auto_state["base_update_id"] = max(int(auto_state.get("base_update_id") or 0), update_id)
                        auto_state["last_check_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        auto_state["deadline_ts"] = int(time.time()) + TELEGRAM_AUTO_TIMEOUT_SECONDS
                        names = auto_state.get("names") or []

                        if is_cancel_message(text_msg):
                            stop_text = build_auto_stop_message("사용자가 '취소'를 보내 자동 실행을 중지했습니다.")
                            auto_state["reply_text"] = text_msg
                            auto_state["status"] = "stopped"
                            auto_state["polling"] = False
                            auto_state["summary_text"] = stop_text
                            st.session_state.bulk_auto_state = auto_state
                            telegram_send_message(stop_text)
                            st.rerun()

                        if status == "waiting_confirm" and is_confirm_message(text_msg):
                            inputs = auto_state.get("pending_inputs") or {}
                            auto_state["reply_text"] = text_msg
                            memo_text = build_auto_memo_text(names, inputs)
                            auto_state["memo_text"] = memo_text
                            telegram_send_message(memo_text)
                            df_products = pd.DataFrame(auto_state.get("products_records") or [])
                            _updated_df, df_changes, df_missing = apply_actions_to_products_df(df_products, cfg, inputs)
                            df_to_push = df_changes.copy() if isinstance(df_changes, pd.DataFrame) else pd.DataFrame()
                            if not df_to_push.empty and "증감" in df_to_push.columns:
                                df_to_push = df_to_push[df_to_push["증감"].astype(float) > 0].copy()
                            api_result = push_stock_updates(df_to_push)
                            summary_text = summarize_auto_result(api_result.get("sent", 0), df_changes, df_missing, api_result)
                            auto_state["summary_text"] = summary_text
                            auto_state["status"] = "done"
                            auto_state["polling"] = False
                            st.session_state.bulk_auto_state = auto_state
                            telegram_send_message(summary_text)
                            st.rerun()

                        ok, parsed_reply, err_text = parse_first_reply_message(text_msg, len(names))
                        if ok:
                            inputs = build_auto_inputs(names, parsed_reply)
                            confirm_text = build_confirmation_message(names, inputs)
                            auto_state["reply_text"] = text_msg
                            auto_state["pending_inputs"] = inputs
                            auto_state["confirm_text"] = confirm_text
                            auto_state["status"] = "waiting_confirm"
                            st.session_state.bulk_auto_state = auto_state
                            telegram_send_message(confirm_text)
                            st.rerun()

                        invalid_text = build_invalid_reply_message(names, err_text)
                        auto_state["reply_text"] = text_msg
                        st.session_state.bulk_auto_state = auto_state
                        telegram_send_message(invalid_text)
                        st.rerun()

                    st.session_state.bulk_auto_state = auto_state
                    time.sleep(TELEGRAM_AUTO_POLL_SECONDS)
                    st.rerun()
                except Exception as e:
                    auto_state["status"] = "error"
                    auto_state["polling"] = False
                    auto_state["error"] = str(e)
                    st.session_state.bulk_auto_state = auto_state
                    try:
                        telegram_send_message(build_auto_stop_message(f"오류로 자동 실행이 중지되었습니다: {e}"))
                    except Exception:
                        pass
                    st.rerun()

            auto_state = st.session_state.get("bulk_auto_state") or {}
            status = auto_state.get("status", "idle")
            poll_active = bool(auto_state.get("polling"))

            if status == "error" and auto_state.get("error"):
                st.error(auto_state.get("error"))
            elif status == "done":
                st.success("자동 반영이 완료되었습니다.")
            elif status == "stopped":
                st.info("자동 실행이 중지되었습니다.")
            elif poll_active:
                st.info("텔레그램 첫 답장 1개를 기다리는 중입니다.")
            else:
                st.caption("자동 실행 시작을 누르면 텔레그램 답장을 자동 확인합니다.")

            if auto_state.get("message_text"):
                st.text(auto_state.get("message_text"))
            if auto_state.get("reply_text"):
                st.caption("처리된 답장")
                st.text(auto_state.get("reply_text"))
            if auto_state.get("memo_text"):
                st.caption("전송된 메모")
                st.text(auto_state.get("memo_text"))
            if auto_state.get("summary_text"):
                st.caption("최종 결과")
                st.text(auto_state.get("summary_text"))
            if auto_state.get("last_check_at"):
                st.caption(f"마지막 확인: {auto_state.get('last_check_at')}")

            if poll_active and status == "waiting_reply":
                try:
                    updates = telegram_get_updates(offset=int(auto_state.get("base_update_id", 0)) + 1)
                    valid_update = None
                    for upd in updates:
                        try:
                            update_id = int(upd.get("update_id") or 0)
                        except Exception:
                            continue
                        msg = upd.get("message") or upd.get("edited_message") or {}
                        chat = msg.get("chat") or {}
                        if str(chat.get("id") or "") != str(TELEGRAM_CHAT_ID):
                            continue
                        if int(msg.get("date") or 0) < int(auto_state.get("sent_at") or 0):
                            continue
                        text_msg = str(msg.get("text") or "").strip()
                        ok, parsed_reply, _err = parse_first_reply_message(text_msg, len(auto_state.get("names") or []))
                        if ok:
                            valid_update = (update_id, text_msg, parsed_reply)
                            break
                    auto_state["last_check_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    if valid_update is not None:
                        update_id, reply_text, parsed_reply = valid_update
                        auto_state["reply_text"] = reply_text
                        auto_state["processed_update_id"] = update_id
                        names = auto_state.get("names") or []
                        inputs = build_auto_inputs(names, parsed_reply)
                        memo_text = build_auto_memo_text(names, inputs)
                        auto_state["memo_text"] = memo_text
                        telegram_send_message(memo_text)
                        df_products = pd.DataFrame(auto_state.get("products_records") or [])
                        _updated_df, df_changes, df_missing = apply_actions_to_products_df(df_products, cfg, inputs)
                        df_to_push = df_changes.copy() if isinstance(df_changes, pd.DataFrame) else pd.DataFrame()
                        if not df_to_push.empty:
                            df_to_push = df_to_push[df_to_push["증감"].astype(float) > 0].copy()
                        api_result = push_stock_updates(df_to_push)
                        summary_text = summarize_auto_result(api_result.get("sent", 0), df_changes, df_missing, api_result)
                        auto_state["summary_text"] = summary_text
                        auto_state["status"] = "done"
                        auto_state["polling"] = False
                        st.session_state.bulk_auto_state = auto_state
                        telegram_send_message(summary_text)
                        st.rerun()
                    st.session_state.bulk_auto_state = auto_state
                    time.sleep(TELEGRAM_AUTO_POLL_SECONDS)
                    st.rerun()
                except Exception as e:
                    auto_state["status"] = "error"
                    auto_state["polling"] = False
                    auto_state["error"] = str(e)
                    st.session_state.bulk_auto_state = auto_state
                    st.rerun()
    elif page.startswith("②"):
        st.subheader("상품목록 추가/삭제/수정")

        prod_list = cfg.get("products", [])
        if prod_list:
            df_prod_raw = pd.DataFrame(prod_list)
        else:
            df_prod_raw = pd.DataFrame(columns=["keyword", "name"])

        # ensure required columns exist
        for _c in ["keyword", "name"]:
            if _c not in df_prod_raw.columns:
                df_prod_raw[_c] = ""

        # ✅ 컬럼 순서: 키워드(패턴) -> 표시될 상품명
        df_prod = df_prod_raw[["keyword", "name"]].rename(
            columns={
                "keyword": "실제 상품명(패턴)",
                "name": "표시될 상품명",
            }
        )

        st.write("• **실제 상품명(패턴)**은 엑셀 '상품명'에서 매칭할 문자열입니다. (비우면 '표시될 상품명'과 동일하게 처리)")
        st.caption("※ 표에서 입력/수정 후 **저장 버튼을 눌러야** 설정이 저장됩니다.")

        with st.form("prod_form", clear_on_submit=False):
            df_prod_edit = st.data_editor(
                df_prod,
                key="bulk_prod_editor",
                use_container_width=True,
                num_rows="dynamic",
                column_config={
                    "실제 상품명(패턴)": st.column_config.TextColumn("실제 상품명(패턴)"),
                    "표시될 상품명": st.column_config.TextColumn("표시될 상품명"),
                },
            )
            save_prod = st.form_submit_button("💾 상품목록 저장", type="primary", use_container_width=True)

        if save_prod:
            cleaned = []
            for _, r in df_prod_edit.iterrows():
                name = str(r.get("표시될 상품명", "")).strip()
                if not name:
                    continue
                kw = str(r.get("실제 상품명(패턴)", "")).strip() or name
                cleaned.append({"name": name, "keyword": kw})

            cfg["products"] = cleaned
            # drop rules of deleted products
            rules = cfg.get("rules", {}) or {}
            rules = {k: v for k, v in rules.items() if k in {p["name"] for p in cleaned}}
            cfg["rules"] = rules

            save_config(cfg)
            st.success("상품목록 저장 완료!")
            st.rerun()

    elif page.startswith("③"):
            st.subheader("규칙 추가/삭제/수정")



            # 🔎 인식로직 관리는 아래쪽(규칙 저장 버튼 아래)에 있습니다.

            prod_names = [p.get("name") for p in cfg.get("products", []) if p.get("name")]
            if not prod_names:
                st.info("먼저 '② 상품목록 관리'에서 상품을 추가하세요.")
            else:
                base = st.selectbox("규칙을 편집할 기준 상품", options=prod_names)

                st.markdown(
                    """
        - **mul(배수)**: `입력수량 × value` 만큼 더함  
        - **map(매핑)**: 입력값별로 딱 정한 값만 더함 (예: `1=2, 2=3`)  
          - map은 **매핑(table)** 칸에 `입력값=적용값`을 `,`로 구분해서 적어요.
        - **키워드 자동 스코프(중요)**: 기준상품을 예) **엔다이브**로 선택한 상태에서 키워드를 `1kg`처럼 **옵션만** 쓰면,  
          엑셀 상품명에 `엔다이브`와 `1kg`가 **둘 다 포함된 행만** 적용됩니다. (다른 상품의 `1kg`는 영향 없음)  
          이미 `엔다이브1kg`처럼 **전체 문자열**을 쓰면 그대로 그 문자열로 매칭합니다.
                    """
                )

                rule_list = (cfg.get("rules", {}) or {}).get(base, [])
                ui_rows = []
                for rr in rule_list:
                    rr2 = Rule.from_dict(rr)
                    table = rr2.table or {}

                    def _sort_key(x):
                        try:
                            return float(x) if x != "*" else 1e19
                        except Exception:
                            return 1e19

                    map_str = ", ".join([f"{k}={fmt_int(table[k])}" for k in sorted(table.keys(), key=_sort_key)]) if table else ""

                    ui_rows.append({
                        "키워드(엑셀 상품명 포함 문자열)": rr2.keyword,
                        "모드": rr2.mode,
                        "value": fmt_int(rr2.value),
                        "매핑(table) - map 모드에서만": map_str,
                    })

                df_rule = pd.DataFrame(ui_rows) if ui_rows else pd.DataFrame(columns=[
                    "키워드(엑셀 상품명 포함 문자열)", "모드", "value", "매핑(table) - map 모드에서만"
                ])

                df_rule_edit = st.data_editor(
                    df_rule,
                    key="bulk_rule_editor",
                    use_container_width=True,
                    num_rows="dynamic",
                    column_config={
                        "키워드(엑셀 상품명 포함 문자열)": st.column_config.TextColumn("키워드"),
                        "모드": st.column_config.SelectboxColumn("모드", options=["mul", "map"]),
                        "value": st.column_config.TextColumn("value", help="mul 모드에서만 사용 (정수)"),
                        "매핑(table) - map 모드에서만": st.column_config.TextColumn(
                            "매핑(table)",
                            help="예: 1=2, 2=3\n(줄바꿈도 가능)",
                        ),
                    },
                )

                if st.button("💾 규칙 저장"):
                    cleaned = []
                    for _, r in df_rule_edit.iterrows():
                        kw = str(r.get("키워드(엑셀 상품명 포함 문자열)", "")).strip()
                        if not kw:
                            continue
                        mode = str(r.get("모드", "mul")).strip()
                        val_raw = r.get("value", 1)
                        if str(val_raw).strip() == "":
                            val = 1
                        else:
                            try:
                                val = int(round(parse_input_number(val_raw)))
                            except Exception:
                                val = 1

                        table_str = str(r.get("매핑(table) - map 모드에서만", "") or "")
                        table = parse_map_string(table_str) if mode == "map" else {}

                        cleaned.append({
                            "keyword": kw,
                            "mode": mode,
                            "value": val,
                            "table": table,
                        })

                    rules = cfg.get("rules", {}) or {}
                    rules[base] = cleaned
                    cfg["rules"] = rules
                    save_config(cfg)
                    st.success("규칙 저장 완료!")
                    st.rerun()


            st.write("")

            # ----------------------------
            # Recognition logic editor (moved from sidebar)
            # ----------------------------
            with st.expander("🔎 인식로직 관리(단위 인식)", expanded=False):
                st.caption("상품명/규칙 키워드에서 단위를 인식하는 로직입니다. (priority는 내부적으로 행 순서로 자동 부여됩니다)")

                logic_rows = _get_recognition_logic(cfg)
                df_logic = pd.DataFrame(
                    [
                        {
                            "priority": int(r.get("priority", 999)),
                            "output_unit": str(r.get("output_unit", "")),
                            "multiplier": str(r.get("multiplier", "1")),
                            "aliases": ", ".join([str(a) for a in (r.get("aliases") or [])]),
                        }
                        for r in logic_rows
                    ]
                )
                if df_logic.empty:
                    df_logic = pd.DataFrame(columns=["priority", "output_unit", "multiplier", "aliases"])

                # 화면에서는 priority를 숨기고, 저장 시 행 순서대로 자동 부여합니다.
                df_ui = df_logic.sort_values("priority").reset_index(drop=True).drop(columns=["priority"], errors="ignore")

                edited = st.data_editor(
                    df_ui,
                    use_container_width=True,
                    num_rows="dynamic",
                    hide_index=True,
                    column_config={
                        "output_unit": st.column_config.TextColumn("output_unit", help="최종 표시 단위 (예: 단, kg, 팩)"),
                        "multiplier": st.column_config.TextColumn("multiplier", help="숫자에 곱해지는 값 (예: g→kg 환산은 0.001)"),
                        "aliases": st.column_config.TextColumn("aliases (쉼표로 구분)", help="인식할 문자열들. 예: kg,킬로,키로"),
                    },
                    key="bulk_recognition_logic_editor",
                )

                c1, c2 = st.columns(2)
                if c1.button("💾 인식로직 저장", use_container_width=True, key="bulk_save_recognition_logic"):
                    raw_rows: List[Dict[str, Any]] = []
                    try:
                        for idx, row in edited.iterrows():
                            unit = str(row.get("output_unit", "")).strip()
                            aliases_raw = str(row.get("aliases", "")).strip()
                            if not unit or not aliases_raw:
                                continue

                            mult = str(row.get("multiplier", "1")).strip() or "1"
                            aliases = [a.strip() for a in aliases_raw.split(",") if a.strip()]

                            # ✅ priority는 화면에 숨기고, 현재 행 순서대로 자동 부여
                            pr = int((idx + 1) * 10)

                            raw_rows.append({"priority": pr, "output_unit": unit, "multiplier": mult, "aliases": aliases})

                        cfg["recognition_logic"] = _normalize_recognition_logic(raw_rows)
                        save_config(cfg)
                        st.success("인식로직 저장 완료!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"저장 실패: {e}")

                # ⬇️ 내보내기: 현재 인식로직을 JSON으로 다운로드
                logic_now = _normalize_recognition_logic(cfg.get("recognition_logic")) or DEFAULT_RECOGNITION_LOGIC.copy()
                export_bytes = json.dumps(logic_now, ensure_ascii=False, indent=2).encode("utf-8")
                export_name = f"recognition_logic_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                c2.download_button("⬇️ 내보내기(JSON)", data=export_bytes, file_name=export_name, mime="application/json", use_container_width=True)

                # 가드: '단' 항목이 없으면 '1단' 키워드도 kg로 떨어질 수 있습니다.
                units_now = [str(r.get("output_unit", "")).strip() for r in logic_now]
                if "단" not in units_now:
                    st.warning("⚠️ 현재 인식로직에 '단' 항목이 없습니다. '1단' 키워드를 써도 kg로 인식될 수 있어요.")

    else:
        st.subheader("설정 백업/복원 (상품목록 + 규칙)")
        cfg_json = json.dumps(cfg, ensure_ascii=False, indent=2).encode("utf-8")
        st.download_button("⬇️ 설정(JSON) 다운로드", data=cfg_json, file_name="stock_config.json", mime="application/json")

        up_cfg = st.file_uploader("설정(JSON) 업로드하여 복원", type=["json"], key="bulk_cfg_uploader")

        # ✅ 복원 완료 알람: 3초만 표시 후 자동으로 사라짐
        if "restore_notice" not in st.session_state:
            st.session_state.restore_notice = None  # ("success"|"error", message)

        restore_clicked = st.button("♻️ 설정 복원", disabled=(up_cfg is None))
        notice_ph = st.empty()

        if restore_clicked:
            try:
                new_cfg = json.loads(up_cfg.getvalue().decode("utf-8"))
                if "products" not in new_cfg or "rules" not in new_cfg:
                    raise ValueError("형식이 올바르지 않습니다.")
                cfg = new_cfg
                save_config(cfg)
                st.session_state.restore_notice = ("success", "설정 복원 완료! 바로 반영됩니다.")
            except Exception as e:
                st.session_state.restore_notice = ("error", f"복원 실패: {e}")

        # 버튼 바로 아래에 복원 결과 표시 (3초 후 자동 삭제)
        if st.session_state.restore_notice:
            kind, msg = st.session_state.restore_notice
            if kind == "success":
                notice_ph.success(msg)
            else:
                notice_ph.error(msg)

            time.sleep(3)
            notice_ph.empty()
            st.session_state.restore_notice = None



# =====================================================
# Router
